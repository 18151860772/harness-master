# -*- coding: utf-8 -*-
"""
Wire Chart 生成工具 V2
- 使用pandas高效处理Excel
- 后台线程处理大数据
- 分页显示避免卡顿
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import threading
import time
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 配置
MAX_ROWS_PER_PAGE = 500  # 每页最大行数


class WireChartTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Wire Chart 生成工具 V2")
        self.root.geometry("1400x800")

        # 数据存储
        self.ecr_wire_df = None
        self.ecr_masterlist_df = None
        self.chart_df = None
        self.part_numbers = []
        self.generated_data = None

        # 分页
        self.current_page = 0
        self.total_pages = 0
        self.filtered_data = []

        self.setup_ui()

    def setup_ui(self):
        """设置UI"""
        # 顶部工具栏
        toolbar = tk.Frame(self.root, bg="#e0e0e0", pady=5)
        toolbar.pack(fill=tk.X, padx=5, pady=5)

        # 文件选择区
        file_frame = tk.Frame(toolbar, bg="#e0e0e0")
        file_frame.pack(side=tk.LEFT, padx=10)

        tk.Button(file_frame, text="上传ECR文件", command=self.load_ecr,
                  bg="#2170b8", fg="white", padx=10).pack(side=tk.LEFT, padx=5)
        self.ecr_label = tk.Label(file_frame, text="未选择", fg="gray", bg="#e0e0e0")
        self.ecr_label.pack(side=tk.LEFT, padx=5)

        tk.Button(file_frame, text="上传Chart文件", command=self.load_chart,
                  bg="#2170b8", fg="white", padx=10).pack(side=tk.LEFT, padx=5)
        self.chart_label = tk.Label(file_frame, text="未选择", fg="gray", bg="#e0e0e0")
        self.chart_label.pack(side=tk.LEFT, padx=5)

        # 生成按钮
        tk.Button(toolbar, text="生成更新", command=self.generate,
                  bg="#28a745", fg="white", padx=15).pack(side=tk.LEFT, padx=20)

        # 导出按钮
        tk.Button(toolbar, text="导出Excel", command=self.export_excel,
                  bg="#28a745", fg="white", padx=15, state=tk.DISABLED).pack(side=tk.LEFT, padx=5)

        # 统计信息
        self.stats_label = tk.Label(toolbar, text="", bg="#e0e0e0", fg="#333")
        self.stats_label.pack(side=tk.LEFT, padx=20)

        # 页面导航
        nav_frame = tk.Frame(toolbar, bg="#e0e0e0")
        nav_frame.pack(side=tk.RIGHT, padx=10)

        tk.Button(nav_frame, text="上一页", command=self.prev_page,
                  width=8).pack(side=tk.LEFT, padx=2)
        self.page_label = tk.Label(nav_frame, text="0/0", bg="#e0e0e0")
        self.page_label.pack(side=tk.LEFT, padx=5)
        tk.Button(nav_frame, text="下一页", command=self.next_page,
                  width=8).pack(side=tk.LEFT, padx=2)

        # 表格区域
        table_frame = tk.Frame(self.root)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 创建Treeview
        self.tree = ttk.Treeview(table_frame, show="headings")
        self.tree.bind("<MouseWheel>", self.on_mousewheel)

        # 滚动条
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # 状态栏
        status_bar = tk.Label(self.root, text="就绪", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)

        # 日志窗口（可折叠）
        self.log_expanded = False
        log_btn = tk.Button(self.root, text="▼ 显示日志", command=self.toggle_log,
                           bg="#f0f0f0", relief=tk.FLAT)
        log_btn.pack(fill=tk.X, side=tk.BOTTOM)

        self.log_text = scrolledtext.ScrolledText(self.root, height=10, state=tk.DISABLED)
        self.log_text.pack(fill=tk.X, side=tk.BOTTOM)

    def log(self, message):
        """添加日志"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"[{time.strftime('%H:%M:%S')}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def toggle_log(self):
        """切换日志显示"""
        if self.log_expanded:
            self.log_text.pack_forget()
            self.log_expanded = False
        else:
            self.log_text.pack(fill=tk.X, side=tk.BOTTOM)
            self.log_expanded = True

    def load_ecr(self):
        """加载ECR文件"""
        filepath = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not filepath:
            return

        self.ecr_path = filepath
        self.ecr_label.config(text=os.path.basename(filepath), fg="#2170b8")
        self.log(f"已选择ECR文件: {filepath}")

        # 后台线程读取
        threading.Thread(target=self._load_ecr_data, daemon=True).start()

    def _load_ecr_data(self):
        """后台读取ECR数据"""
        try:
            self.log("正在读取ECR文件...")
            wb = load_workbook(self.ecr_path, read_only=True, data_only=True)

            # 读取WIRE sheet
            wire_sheet_name = None
            for name in wb.sheetnames:
                if 'WIRE' in name:
                    wire_sheet_name = name
                    break

            masterlist_sheet_name = None
            for name in wb.sheetnames:
                if 'Masterlist' in name and 'CHG' not in name:
                    masterlist_sheet_name = name
                    break

            if wire_sheet_name:
                self.log(f"找到WIRE sheet: {wire_sheet_name}")
                self.ecr_wire_df = pd.read_excel(wb, sheet_name=wire_sheet_name, header=None)
                self.log(f"WIRE数据: {len(self.ecr_wire_df)} 行")

            if masterlist_sheet_name:
                self.log(f"找到Masterlist sheet: {masterlist_sheet_name}")
                self.ecr_masterlist_df = pd.read_excel(wb, sheet_name=masterlist_sheet_name, header=None)
                self.log(f"Masterlist数据: {len(self.ecr_masterlist_df)} 行")

            wb.close()
            self.log("ECR文件读取完成")

        except Exception as e:
            self.log(f"读取ECR出错: {e}")

    def load_chart(self):
        """加载Chart文件"""
        filepath = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not filepath:
            return

        self.chart_path = filepath
        self.chart_label.config(text=os.path.basename(filepath), fg="#2170b8")
        self.log(f"已选择Chart文件: {filepath}")

        # 后台线程读取
        threading.Thread(target=self._load_chart_data, daemon=True).start()

    def _load_chart_data(self):
        """后台读取Chart数据"""
        try:
            self.log("正在读取Chart文件...")
            wb = load_workbook(self.chart_path, read_only=True, data_only=True)

            wire_chart_sheet_name = None
            for name in wb.sheetnames:
                if 'Wire Chart' in name:
                    wire_chart_sheet_name = name
                    break

            if wire_chart_sheet_name:
                self.log(f"找到Wire Chart sheet: {wire_chart_sheet_name}")
                self.chart_df = pd.read_excel(wb, sheet_name=wire_chart_sheet_name, header=None)
                self.log(f"Wire Chart数据: {len(self.chart_df)} 行")

                # 从第3行（索引2）开始读取零件号
                if len(self.chart_df) > 2:
                    header_row = self.chart_df.iloc[2].tolist()
                    self.part_numbers = []
                    for i in range(28, len(header_row)):
                        val = header_row[i]
                        if pd.notna(val) and val != 'option' and str(val).strip():
                            self.part_numbers.append(str(val).strip())
                    self.log(f"自动读取到 {len(self.part_numbers)} 个零件号")

            wb.close()
            self.log("Chart文件读取完成")

        except Exception as e:
            self.log(f"读取Chart出错: {e}")

    def generate(self):
        """生成更新数据"""
        if self.ecr_wire_df is None or self.chart_df is None:
            messagebox.showwarning("警告", "请先上传ECR和Chart文件")
            return

        if self.ecr_masterlist_df is None:
            messagebox.showwarning("警告", "未找到Masterlist sheet")
            return

        # 后台线程处理
        threading.Thread(target=self._generate_data, daemon=True).start()

    def _generate_data(self):
        """后台生成数据"""
        try:
            self.log("开始生成Wire Chart...")

            # 构建WIRE映射 (B列=索引1是Wire ID)
            self.log("构建WIRE数据映射...")
            wire_map = {}
            for idx, row in self.ecr_wire_df.iterrows():
                if idx >= 2 and pd.notna(row[1]):  # B列是Wire ID
                    wire_id = str(row[1]).strip()
                    wire_map[wire_id] = row

            self.log(f"WIRE映射条目: {len(wire_map)}")

            # 构建零件号配置缓存
            self.log("构建零件号配置...")
            masterlist_header = self.ecr_masterlist_df.iloc[0].tolist()
            part_config_cache = {}

            for part_num in self.part_numbers:
                configs = set()
                for idx, row in self.ecr_masterlist_df.iterrows():
                    if idx >= 2 and pd.notna(row[0]):
                        if str(row[0]).strip() == part_num:
                            for j in range(8, len(masterlist_header)):
                                val = row[j]
                                if pd.notna(val) and str(val).strip() and str(val).strip() != '-':
                                    configs.add(str(val).strip())
                            break
                part_config_cache[part_num] = list(configs)

            self.log(f"零件号配置缓存: {len(part_config_cache)}")

            # 列映射
            col_map = {
                2: 1,   # CIRCUIT NBR -> Wire ID
                3: 3,   # WIRE SIZE -> Size / Gauge
                6: 2,   # COLOR -> Color
                8: 4,   # CABLE DESG -> Material
                9: 7,   # EST WIRE LENGTH -> Ident Tag
                10: 6,  # GROUP NAME -> Multicore ID
                11: 8,  # POS NBR 1 -> From Code
                12: 9,  # CAV 1 -> From Pin
                21: 10, # POS NBR 2 -> To Code
                22: 11  # CAV 2 -> To Pin
            }
            wire_option_col = 5  # F列是Option

            # 查找零件号在Chart中的列位置
            chart_header = self.chart_df.iloc[2].tolist()
            part_col_map = {}
            for pn in self.part_numbers:
                for i, val in enumerate(chart_header):
                    if pd.notna(val) and str(val).strip() == pn:
                        part_col_map[pn] = i
                        break

            # 处理数据
            self.log("处理回路数据...")
            generated_rows = []
            chart_rows = self.chart_df.iloc[3:]  # 跳过前3行header

            for idx, chart_row in chart_rows.iterrows():
                circuit_nbr = chart_row[2]
                if pd.isna(circuit_nbr):
                    continue

                wire_id = str(circuit_nbr).strip()
                wire_row = wire_map.get(wire_id)

                if wire_row is not None:
                    new_row = list(chart_row)
                    updated_cols = []

                    # 更新基本信息
                    for chart_col, wire_col in col_map.items():
                        wire_val = wire_row[wire_col]
                        if pd.notna(wire_val) and str(wire_val).strip():
                            new_row[chart_col] = str(wire_val).strip()
                            updated_cols.append(chart_col)

                    # 计算X打点
                    option_expr = wire_row[wire_option_col]
                    if pd.notna(option_expr):
                        option_str = str(option_expr).strip()

                        # 清除现有打点
                        for col_idx in part_col_map.values():
                            new_row[col_idx] = ''

                        # 计算打点
                        for pn in self.part_numbers:
                            part_configs = part_config_cache.get(pn, [])
                            if self._calculate_option(option_str, part_configs):
                                part_col_idx = part_col_map.get(pn)
                                if part_col_idx is not None:
                                    new_row[part_col_idx] = 'X'

                    new_row._updated_cols = updated_cols
                    generated_rows.append(new_row)

            self.generated_data = generated_rows
            self.filtered_data = generated_rows
            self.current_page = 0
            self.total_pages = max(1, (len(self.filtered_data) + MAX_ROWS_PER_PAGE - 1) // MAX_ROWS_PER_PAGE)

            self.log(f"生成完成: {len(generated_rows)} 行数据")
            self.log(f"页面数: {self.total_pages}")

            # 更新UI
            self.root.after(0, self._show_results)

        except Exception as e:
            self.log(f"生成出错: {e}")
            import traceback
            self.log(traceback.format_exc())

    def _calculate_option(self, expression, car_configs):
        """计算option表达式"""
        if not expression or expression == 'NaN':
            return False

        expr = expression.replace(/\s+/g, '').upper()
        tokens = [t for t in expr.split(/([&\/\-\(\)])/) if t]

        def eval_tokens(tokens, start, end):
            result = None
            current_op = None
            i = start

            while i < end:
                token = tokens[i]

                if token == '(':
                    depth = 1
                    j = i + 1
                    while j < end and depth > 0:
                        if tokens[j] == '(':
                            depth += 1
                        elif tokens[j] == ')':
                            depth -= 1
                        j += 1
                    value = eval_tokens(tokens, i + 1, j - 1)
                    i = j
                elif token == '-':
                    i += 1
                    if i >= end:
                        break
                    next_token = tokens[i]
                    if next_token == '(':
                        depth = 1
                        j = i + 1
                        while j < end and depth > 0:
                            if tokens[j] == '(':
                                depth += 1
                            elif tokens[j] == ')':
                                depth -= 1
                            j += 1
                        value = eval_tokens(tokens, i + 1, j - 1)
                        value = not any(value) if value else False
                        i = j
                    else:
                        value = next_token in car_configs
                        i += 1
                elif token in ('&', '/', ''):
                    i += 1
                    continue
                elif token == ')':
                    break
                else:
                    value = token in car_configs
                    i += 1

                if result is None:
                    result = [value]
                elif current_op == '&':
                    result = [r and value for r, value in zip(result, [value] * len(result))]
                elif current_op == '/':
                    result = [r or value for r, value in zip(result, [value] * len(result))]

                if i < end and tokens[i] in ('&', '/'):
                    current_op = tokens[i]
                    i += 1

            return result

        result = eval_tokens(tokens, 0, len(tokens))
        return result and result[0]

    def _show_results(self):
        """显示结果"""
        # 更新统计
        x_count = sum(1 for row in self.generated_data for cell in row if cell == 'X')
        self.stats_label.config(
            text=f"回路: {len(self.generated_data)} | 零件号: {len(self.part_numbers)} | X打点: {x_count}"
        )

        # 启用导出按钮
        for child in self.root.winfo_children():
            if isinstance(child, tk.Frame):
                for btn in child.winfo_children():
                    if isinstance(btn, tk.Button) and btn.cget("text") == "导出Excel":
                        btn.config(state=tk.NORMAL)
                        break

        self.log("显示结果...")
        self.show_page()

    def show_page(self):
        """显示当前页"""
        if not self.filtered_data:
            return

        start_idx = self.current_page * MAX_ROWS_PER_PAGE
        end_idx = min(start_idx + MAX_ROWS_PER_PAGE, len(self.filtered_data))
        page_data = self.filtered_data[start_idx:end_idx]

        # 获取表头
        if len(self.chart_df) >= 3:
            header_row = self.chart_df.iloc[2].tolist()
        else:
            header_row = list(range(len(self.chart_df.columns)))

        # 显示列（基本信息 + 零件号）
        display_cols = [0, 1, 2, 3, 6, 8, 9, 10, 11, 12, 21, 22]
        part_col_indices = []
        for i, val in enumerate(header_row):
            if pd.notna(val) and str(val).strip() in self.part_numbers:
                part_col_indices.append(i)

        all_cols = list(dict.fromkeys(display_cols + part_col_indices))
        all_cols.sort()

        # 设置列
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = all_cols

        for col in all_cols:
            header = header_row[col] if pd.notna(header_row[col]) else str(col)
            self.tree.heading(col, text=str(header))
            self.tree.column(col, width=80, minwidth=50)

        # 插入数据
        for row_idx, row in enumerate(page_data):
            values = []
            for col in all_cols:
                val = row[col] if col < len(row) else ''
                if pd.notna(val):
                    values.append(str(val))
                else:
                    values.append('')
            self.tree.insert("", tk.END, values=values, iid=str(row_idx))

        # 更新页码
        self.page_label.config(text=f"{self.current_page + 1}/{self.total_pages}")
        self.log(f"显示第 {self.current_page + 1} 页 ({start_idx + 1}-{end_idx})")

    def prev_page(self):
        """上一页"""
        if self.current_page > 0:
            self.current_page -= 1
            self.show_page()

    def next_page(self):
        """下一页"""
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.show_page()

    def export_excel(self):
        """导出Excel"""
        if not self.generated_data:
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Wire_Chart_更新.xlsx"
        )
        if not filepath:
            return

        self.log(f"正在导出: {filepath}")

        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Wire Chart"

            # 写入表头
            if len(self.chart_df) >= 3:
                header_row = self.chart_df.iloc[2].tolist()
            else:
                header_row = list(range(len(self.chart_df.columns)))

            for col_idx, val in enumerate(header_row):
                cell = ws.cell(row=1, column=col_idx + 1)
                if pd.notna(val):
                    cell.value = str(val)

            # 写入数据
            for row_idx, row in enumerate(self.generated_data):
                for col_idx, val in enumerate(row):
                    cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
                    if pd.notna(val):
                        cell.value = str(val)

            # 调整列宽
            for col_idx in range(len(header_row)):
                col_letter = get_column_letter(col_idx + 1)
                ws.column_dimensions[col_letter].width = 12

            wb.save(filepath)
            self.log(f"导出完成: {filepath}")
            messagebox.showinfo("完成", "导出成功！")

        except Exception as e:
            self.log(f"导出出错: {e}")
            messagebox.showerror("错误", f"导出失败: {e}")

    def on_mousewheel(self, event):
        """鼠标滚轮事件"""
        self.tree.yview_scroll(int(-1 * (event.delta / 120)), "units")


def main():
    root = tk.Tk()
    app = WireChartTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()
