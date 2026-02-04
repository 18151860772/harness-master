# -*- coding: utf-8 -*-
"""
Wire Chart 导出工具（带格式）
用法: python export_with_format.py <Chart文件> <输出文件> <JSON数据文件>
"""

import sys
import json
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def main():
    if len(sys.argv) < 4:
        print("用法: python export_with_format.py <Chart文件> <输出文件> <JSON数据文件>")
        return

    chart_file = sys.argv[1]
    output_file = sys.argv[2]
    json_file = sys.argv[3]

    print(f"Chart文件: {chart_file}")
    print(f"输出文件: {output_file}")
    print(f"JSON数据: {json_file}")

    # 读取JSON数据
    print("正在读取JSON数据...")
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    chartHeader = data.get('chartHeader', [])
    dataRows = data.get('dataRows', [])
    partNumbers = data.get('partNumbers', [])

    print(f"表头长度: {len(chartHeader)}")
    print(f"数据行数: {len(dataRows)}")
    print(f"零件号数: {len(partNumbers)}")

    if not chart_file or not os.path.exists(chart_file):
        # 没有原Chart文件，创建新文件
        print("未提供原Chart文件，创建新文件...")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Wire Chart"

        # 写入表头
        for col_idx, val in enumerate(chartHeader, start=1):
            ws.cell(row=1, column=col_idx, value=val if val else None)
            ws.cell(row=2, column=col_idx, value=None)
            ws.cell(row=3, column=col_idx, value=None)
            ws.cell(row=4, column=col_idx, value=None)

        # 写入数据
        for row_idx, row in enumerate(dataRows, start=5):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val if val else None)

        wb.save(output_file)
        print(f"导出完成: {output_file}")
        return

    # 读取原Chart文件格式
    print("正在读取Chart格式...")
    source_wb = load_workbook(chart_file)
    source_ws = None
    for name in source_wb.sheetnames:
        if 'Wire Chart' in name:
            source_ws = source_wb[name]
            break

    if not source_ws:
        print("错误: 未找到Wire Chart sheet")
        return

    # 复制格式到新工作簿
    print("正在复制格式...")
    from openpyxl import Workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Wire Chart"

    # 获取原工作表的最大列数
    max_col = max(source_ws.max_column, len(chartHeader))

    # 复制前4行格式
    for row_idx in range(1, 5):
        for col_idx in range(1, max_col + 1):
            cell = source_ws.cell(row=row_idx, column=col_idx)
            new_cell = out_ws.cell(row=row_idx, column=col_idx, value=cell.value)
            if cell.value is None:
                continue
            if cell.has_style:
                try:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.alignment = cell.alignment.copy()
                    new_cell.number_format = cell.number_format
                except:
                    pass

    # 复制列宽
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in source_ws.column_dimensions:
            dim = source_ws.column_dimensions[col_letter]
            if dim.width:
                out_ws.column_dimensions[col_letter].width = dim.width

    # 复制行高
    for row_idx in range(1, 5):
        if row_idx in source_ws.row_dimensions:
            height = source_ws.row_dimensions[row_idx].height
            if height:
                out_ws.row_dimensions[row_idx].height = height

    # 写入新数据（从第5行开始）
    print("正在写入数据...")
    for row_idx, row in enumerate(dataRows, start=5):
        for col_idx, val in enumerate(row, start=1):
            cell = out_ws.cell(row=row_idx, column=col_idx, value=val if val else None)

            # 尝试从原工作表获取对应位置的样式
            try:
                source_row = source_ws.max_row
                if row_idx - 4 <= source_row:
                    source_cell = source_ws.cell(row=row_idx, column=col_idx)
                    if source_cell.has_style:
                        cell.font = source_cell.font.copy()
                        cell.fill = source_cell.fill.copy()
                        cell.border = source_cell.border.copy()
                        cell.alignment = source_cell.alignment.copy()
            except:
                pass

    out_wb.save(output_file)
    print(f"导出完成: {output_file}")


if __name__ == "__main__":
    main()
