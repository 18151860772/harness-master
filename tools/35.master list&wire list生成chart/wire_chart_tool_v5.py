# -*- coding: utf-8 -*-
"""
Wire Chart 生成工具 V5
- 保留原chart文件的颜色格式
- 完整复制前4行格式
- 从WIRE提取回路数据
- 根据Masterlist计算option打点
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import threading
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class WireChartToolV5:
    def __init__(self, root):
        self.root = root
        self.root.title("Wire Chart 生成工具 V5")
        self.root.geometry("1400x800")

        # 数据存储
        self.wire_df = None
        self.masterlist_df = None
        self.part_numbers = []
        self.generated_data = None
        self.chart_file_path = None
        self.chart_wb = None  # 原chart workbook（用于复制格式）

        # 分页
        self.current_page = 0
        self.total_pages = 0
        self.page_size = 500

        self.setup_ui()

    def setup_ui(self):
        """设置UI"""
        # 顶部工具栏
        toolbar = tk.Frame(self.root, bg="#e0e0e0", pady=5)
        toolbar.pack(fill=tk.X, padx=5, pady=5)

        # 文件选择区
        file_frame = tk.Frame(toolbar, bg="#e0e0e0")
        file_frame.pack(side=tk.LEFT, padx=10)

        tk.Button(file_frame, text="上传ECR文件", command=self.load_ecr,
                  bg="#2170b8", fg="white", padx=10).pack(side=tk.LEFT, padx=5)
        self.ecr_label = tk.Label(file_frame, text="未选择", fg="gray", bg="#e0e0e0")
        self.ecr_label.pack(side=tk.LEFT, padx=5)

        tk.Button(file_frame, text="上传Chart文件", command=self.load_chart,
                  bg="#2170b8", fg="white", padx=10).pack(side=tk.LEFT, padx=5)
        self.chart_label = tk.Label(file_frame, text="未选择", fg="gray", bg="#e0e0e0")
        self.chart_label.pack(side=tk.LEFT, padx=5)

        # 生成按钮
        tk.Button(toolbar, text="生成Wire Chart", command=self.generate,
                  bg="#28a745", fg="white", padx=15).pack(side=tk.LEFT, padx=20)

        # 导出按钮
        tk.Button(toolbar, text="导出Excel", command=self.export_excel,
                  bg="#28a745", fg="white", padx=15, state=tk.DISABLED).pack(side=tk.LEFT, padx=5)

        # 统计信息
        self.stats_label = tk.Label(toolbar, text="", bg="#e0e0e0", fg="#333")
        self.stats_label.pack(side=tk.LEFT, padx=20)

        # 页面导航
        nav_frame = tk.Frame(toolbar, bg="#e0e0e0")
        nav_frame.pack(side=tk.RIGHT, padx=10)

        tk.Button(nav_frame, text="上一页", command=self.prev_page,
                  width=8).pack(side=tk.LEFT, padx=2)
        self.page_label = tk.Label(nav_frame, text="0/0", bg="#e0e0e0")
        self.page_label.pack(side=tk.LEFT, padx=5)
        tk.Button(nav_frame, text="下一页", command=self.next_page,
                  width=8).pack(side=tk.LEFT, padx=2)

        # 表格区域
        table_frame = tk.Frame(self.root)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 创建Treeview
        self.tree = ttk.Treeview(table_frame, show="headings")
        self.tree.bind("<MouseWheel>", self.on_mousewheel)

        # 滚动条
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # 状态栏
        status_bar = tk.Label(self.root, text="就绪", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)

        # 日志窗口
        self.log_expanded = False
        log_btn = tk.Button(self.root, text="▼ 显示日志", command=self.toggle_log,
                           bg="#f0f0f0", relief=tk.FLAT)
        log_btn.pack(fill=tk.X, side=tk.BOTTOM)

        self.log_text = scrolledtext.ScrolledText(self.root, height=10, state=tk.DISABLED)
        self.log_text.pack(fill=tk.X, side=tk.BOTTOM)

    def log(self, message):
        """添加日志"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"[{time.strftime('%H:%M:%S')}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def toggle_log(self):
        """切换日志显示"""
        if self.log_expanded:
            self.log_text.pack_forget()
            self.log_expanded = False
        else:
            self.log_text.pack(fill=tk.X, side=tk.BOTTOM)
            self.log_expanded = True

    def load_ecr(self):
        """加载ECR文件"""
        filepath = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not filepath:
            return

        self.ecr_path = filepath
        self.ecr_label.config(text=os.path.basename(filepath), fg="#2170b8")
        self.log(f"已选择ECR文件: {filepath}")

        threading.Thread(target=self._load_ecr_data, daemon=True).start()

    def _load_ecr_data(self):
        """后台读取ECR数据"""
        try:
            self.log("正在读取ECR文件...")
            wb = load_workbook(self.ecr_path, read_only=True, data_only=True)

            # 读取WIRE sheet
            wire_sheet_name = None
            for name in wb.sheetnames:
                if 'WIRE' in name:
                    wire_sheet_name = name
                    break

            # 读取Masterlist sheet
            masterlist_sheet_name = None
            for name in wb.sheetnames:
                if 'Masterlist' in name and 'CHG' not in name:
                    masterlist_sheet_name = name
                    break

            if wire_sheet_name:
                self.log(f"找到WIRE sheet: {wire_sheet_name}")
                self.wire_df = pd.read_excel(wb, sheet_name=wire_sheet_name, header=None)
                self.log(f"WIRE数据: {len(self.wire_df)} 行")

            if masterlist_sheet_name:
                self.log(f"找到Masterlist sheet: {masterlist_sheet_name}")
                self.masterlist_df = pd.read_excel(wb, sheet_name=masterlist_sheet_name, header=None)
                self.log(f"Masterlist数据: {len(self.masterlist_df)} 行")

            wb.close()
            self.log("ECR文件读取完成")

        except Exception as e:
            self.log(f"读取ECR出错: {e}")

    def load_chart(self):
        """加载Chart文件（保留格式）"""
        filepath = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not filepath:
            return

        self.chart_file_path = filepath
        self.chart_label.config(text=os.path.basename(filepath), fg="#2170b8")
        self.log(f"已选择Chart文件: {filepath}")

        # 加载原文件用于复制格式
        try:
            self.chart_wb = load_workbook(filepath)
            # 查找Wire Chart sheet
            wire_chart_sheet = None
            for name in self.chart_wb.sheetnames:
                if 'Wire Chart' in name:
                    wire_chart_sheet = self.chart_wb[name]
                    break

            if wire_chart_sheet:
                self.log(f"原Chart Sheet: {wire_chart_sheet.title}")
                # 读取表头行
                header_row = []
                for cell in wire_chart_sheet[4]:  # 第4行是表头
                    header_row.append(cell.value)
                self.log(f"表头前10列: {header_row[:10]}")

                # 查找零件号列（从第28列开始）
                part_col_start = 28
                self.part_numbers = []
                for i, val in enumerate(header_row[part_col_start:], start=part_col_start):
                    if val and str(val).strip():
                        self.part_numbers.append(str(val).strip())
                self.log(f"找到 {len(self.part_numbers)} 个零件号")

        except Exception as e:
            self.log(f"读取Chart格式出错: {e}")

    def generate(self):
        """生成Wire Chart"""
        if self.wire_df is None:
            messagebox.showwarning("警告", "请先上传ECR文件")
            return

        if self.masterlist_df is None:
            messagebox.showwarning("警告", "未找到Masterlist sheet")
            return

        threading.Thread(target=self._generate_data, daemon=True).start()

    def _calculate_option(self, expression, car_configs):
        """计算option表达式"""
        if not expression or pd.isna(expression) or str(expression) == 'NaN':
            return False

        expr = str(expression).replace(/\s+/g, '').upper()
        tokens = [t for t in expr.split(/([&/\-\(\)])/) if t]

        def eval_tokens(tokens, start, end):
            result = None
            current_op = None
            i = start

            while i < end:
                token = tokens[i]

                if token == '(':
                    depth = 1
                    j = i + 1
                    while j < end and depth > 0:
                        if tokens[j] == '(':
                            depth += 1
                        elif tokens[j] == ')':
                            depth -= 1
                        j += 1
                    value = eval_tokens(tokens, i + 1, j - 1)
                    i = j
                elif token == '-':
                    i += 1
                    if i >= end:
                        break
                    next_token = tokens[i]
                    if next_token == '(':
                        depth = 1
                        j = i + 1
                        while j < end and depth > 0:
                            if tokens[j] == '(':
                                depth += 1
                            elif tokens[j] == ')':
                                depth -= 1
                            j += 1
                        value = eval_tokens(tokens, i + 1, j - 1)
                        value = not any(value) if value else False
                        i = j
                    else:
                        value = next_token in car_configs
                        i += 1
                elif token in ('&', '/', ''):
                    i += 1
                    continue
                elif token == ')':
                    break
                else:
                    value = token in car_configs
                    i += 1

                if result is None:
                    result = [value]
                elif current_op == '&':
                    result = [r and value for r, value in zip(result, [value] * len(result))]
                elif current_op == '/':
                    result = [r or value for r, value in zip(result, [value] * len(result))]

                if i < end and tokens[i] in ('&', '/'):
                    current_op = tokens[i]
                    i += 1

            return result

        result = eval_tokens(tokens, 0, len(tokens))
        return result and result[0]

    def _generate_data(self):
        """后台生成数据"""
        try:
            self.log("开始生成Wire Chart...")

            # WIRE: A=0, B=1(Wire ID), C=2(Color), D=3(Size/Gauge), E=4(Material), F=5(Option),
            #       G=6(Multicore ID), H=7(Ident Tag), I=8(From Code), J=9(From Pin),
            #       K=10(To Code), L=11(To Pin)
            # 表头在第2行(索引1)，数据从第4行(索引3)开始

            # 构建零件号配置缓存
            self.log("构建零件号配置...")
            masterlist_header = self.masterlist_df.iloc[0].tolist()
            part_config_cache = {}

            for part_num in self.part_numbers:
                configs = set()
                for idx, row in self.masterlist_df.iterrows():
                    if idx >= 2 and pd.notna(row[0]):
                        if str(row[0]).strip() == part_num:
                            for j in range(8, len(masterlist_header)):
                                val = row[j]
                                if pd.notna(val) and str(val).strip() and str(val).strip() != '-':
                                    configs.add(str(val).strip())
                            break
                part_config_cache[part_num] = list(configs)

            self.log(f"零件号配置缓存: {len(part_config_cache)}")

            # 查找表头中各列的位置
            if self.chart_wb:
                # 从原chart获取表头
                wire_chart_sheet = None
                for name in self.chart_wb.sheetnames:
                    if 'Wire Chart' in name:
                        wire_chart_sheet = self.chart_wb[name]
                        break

                header_row = []
                for cell in wire_chart_sheet[4]:
                    header_row.append(cell.value)

                # 查找各列索引
                col_idx_map = {
                    'CIRCUIT_NBR': header_row.index('CIRCUIT NBR') if 'CIRCUIT NBR' in header_row else 2,
                    'WIRE_SIZE': header_row.index('WIRE SIZE') if 'WIRE SIZE' in header_row else 3,
                    'COLOR': header_row.index('COLOR') if 'COLOR' in header_row else 6,
                    'CABLE_DESG': header_row.index('CABLE DESG') if 'CABLE DESG' in header_row else 8,
                    'EST_WIRE_LENGTH': header_row.index('EST WIRE LENGTH') if 'EST WIRE LENGTH' in header_row else 9,
                    'GROUP_NAME': header_row.index('GROUP NAME') if 'GROUP NAME' in header_row else 10,
                    'POS_NBR_1': header_row.index('POS NBR 1') if 'POS NBR 1' in header_row else 11,
                    'CAV_1': header_row.index('CAV 1') if 'CAV 1' in header_row else 12,
                    'OPTION': header_row.index('option') if 'option' in header_row else 22,
                    'POS_NBR_2': header_row.index('POS NBR 2') if 'POS NBR 2' in header_row else 25,
                    'CAV_2': header_row.index('CAV 2') if 'CAV 2' in header_row else 26,
                }
                self.log(f"列索引映射: {col_idx_map}")
            else:
                # 默认列索引
                col_idx_map = {
                    'CIRCUIT_NBR': 2, 'WIRE_SIZE': 3, 'COLOR': 6, 'CABLE_DESG': 8,
                    'EST_WIRE_LENGTH': 9, 'GROUP_NAME': 10, 'POS_NBR_1': 11, 'CAV_1': 12,
                    'OPTION': 22, 'POS_NBR_2': 25, 'CAV_2': 26,
                }

            # 零件号列起始位置
            part_col_start = 28

            # 生成数据
            self.log("生成回路数据...")
            generated_rows = []
            x_count = 0

            for idx in range(3, len(self.wire_df)):
                wire_row = self.wire_df.iloc[idx]
                wire_id = wire_row[1]  # B列是Wire ID

                if pd.isna(wire_id):
                    continue

                new_row = [''] * len(header_row) if self.chart_wb else [''] * 28
                updated_cols = []

                # 填充各列
                new_row[col_idx_map['CIRCUIT_NBR']] = str(wire_row[1]).strip()
                updated_cols.append(col_idx_map['CIRCUIT_NBR'])

                new_row[col_idx_map['WIRE_SIZE']] = str(wire_row[3]).strip() if pd.notna(wire_row[3]) else ''
                updated_cols.append(col_idx_map['WIRE_SIZE'])

                new_row[col_idx_map['COLOR']] = str(wire_row[2]).strip() if pd.notna(wire_row[2]) else ''
                updated_cols.append(col_idx_map['COLOR'])

                new_row[col_idx_map['CABLE_DESG']] = str(wire_row[4]).strip() if pd.notna(wire_row[4]) else ''
                updated_cols.append(col_idx_map['CABLE_DESG'])

                new_row[col_idx_map['EST_WIRE_LENGTH']] = str(wire_row[7]).strip() if pd.notna(wire_row[7]) else ''
                updated_cols.append(col_idx_map['EST_WIRE_LENGTH'])

                new_row[col_idx_map['GROUP_NAME']] = str(wire_row[6]).strip() if pd.notna(wire_row[6]) else ''
                updated_cols.append(col_idx_map['GROUP_NAME'])

                new_row[col_idx_map['POS_NBR_1']] = str(wire_row[8]).strip() if pd.notna(wire_row[8]) else ''
                updated_cols.append(col_idx_map['POS_NBR_1'])

                new_row[col_idx_map['CAV_1']] = str(wire_row[9]).strip() if pd.notna(wire_row[9]) else ''
                updated_cols.append(col_idx_map['CAV_1'])

                new_row[col_idx_map['POS_NBR_2']] = str(wire_row[10]).strip() if pd.notna(wire_row[10]) else ''
                updated_cols.append(col_idx_map['POS_NBR_2'])

                new_row[col_idx_map['CAV_2']] = str(wire_row[11]).strip() if pd.notna(wire_row[11]) else ''
                updated_cols.append(col_idx_map['CAV_2'])

                new_row[col_idx_map['OPTION']] = str(wire_row[5]).strip() if pd.notna(wire_row[5]) else ''

                # 计算X打点
                option_expr = wire_row[5]
                for pi, pn in enumerate(self.part_numbers):
                    part_configs = part_config_cache.get(pn, [])
                    if self._calculate_option(option_expr, part_configs):
                        col_idx = part_col_start + pi
                        if col_idx < len(new_row):
                            new_row[col_idx] = 'X'
                            x_count += 1

                new_row._updated_cols = updated_cols
                generated_rows.append(new_row)

                if len(generated_rows) % 500 == 0:
                    self.log(f"已处理 {len(generated_rows)} 行...")

            self.generated_data = generated_rows
            self.current_page = 0
            self.total_pages = max(1, (len(self.generated_data) + self.page_size - 1) // self.page_size)

            self.log(f"生成完成: {len(generated_rows)} 行数据, X打点: {x_count}")

            self.root.after(0, self._show_results)

        except Exception as e:
            self.log(f"生成出错: {e}")
            import traceback
            self.log(traceback.format_exc())

    def _show_results(self):
        """显示结果"""
        x_count = sum(1 for row in self.generated_data for cell in row if cell == 'X')
        self.stats_label.config(
            text=f"回路: {len(self.generated_data)} | 零件号: {len(self.part_numbers)} | X打点: {x_count}"
        )

        # 启用导出按钮
        for child in self.root.winfo_children():
            if isinstance(child, tk.Frame):
                for btn in child.winfo_children():
                    if isinstance(btn, tk.Button) and btn.cget("text") == "导出Excel":
                        btn.config(state=tk.NORMAL)
                        break

        self.show_page()

    def show_page(self):
        """显示当前页"""
        if not self.generated_data:
            return

        start_idx = self.current_page * self.page_size
        end_idx = min(start_idx + self.page_size, len(self.generated_data))
        page_data = self.generated_data[start_idx:end_idx]

        # 获取表头
        if self.chart_wb:
            wire_chart_sheet = None
            for name in self.chart_wb.sheetnames:
                if 'Wire Chart' in name:
                    wire_chart_sheet = self.chart_wb[name]
                    break
            header_row = [cell.value for cell in wire_chart_sheet[4]]
        else:
            header_row = ['ACTION', 'CHANGE DATE', 'CIRCUIT NBR', 'WIRE SIZE',
                          'Luxshare CABLE PART NBR', 'CUSTOMER CABLE PART NBR', 'COLOR',
                          'MARKING COLOR', 'CABLE DESG', 'EST WIRE LENGTH', 'GROUP NAME',
                          'POS NBR 1', 'CAV 1'] + [''] * (len(self.generated_data[0]) - 13) if self.generated_data else []

        # 设置列
        self.tree.delete(*self.tree.get_children())
        display_cols = list(range(len(header_row)))
        self.tree["columns"] = display_cols

        for col in display_cols:
            header = header_row[col] if col < len(header_row) else str(col)
            self.tree.heading(col, text=str(header))
            self.tree.column(col, width=80, minwidth=50)

        # 插入数据
        for row_idx, row in enumerate(page_data):
            values = []
            for col in display_cols:
                val = row[col] if col < len(row) else ''
                if pd.notna(val):
                    values.append(str(val))
                else:
                    values.append('')
            self.tree.insert("", tk.END, values=values, iid=str(row_idx))

        self.page_label.config(text=f"{self.current_page + 1}/{self.total_pages}")

    def prev_page(self):
        """上一页"""
        if self.current_page > 0:
            self.current_page -= 1
            self.show_page()

    def next_page(self):
        """下一页"""
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.show_page()

    def export_excel(self):
        """导出Excel（保留原格式）"""
        if not self.generated_data:
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Wire_Chart_更新.xlsx"
        )
        if not filepath:
            return

        self.log(f"正在导出: {filepath}")

        try:
            if self.chart_wb:
                # 复制原文件的格式
                source_wb = load_workbook(self.chart_file_path)
                source_ws = None
                for name in source_wb.sheetnames:
                    if 'Wire Chart' in name:
                        source_ws = source_wb[name]
                        break

                # 创建新workbook
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Wire Chart"

                # 复制前4行（格式）
                for row_idx in range(1, 5):  # 1-4行
                    for col_idx, cell in enumerate(source_ws[row_idx], start=1):
                        new_cell = ws.cell(row=row_idx, column=col_idx, value=cell.value)

                        # 复制样式
                        if cell.has_style:
                            new_cell.font = cell.font.copy()
                            new_cell.fill = cell.fill.copy()
                            new_cell.border = cell.border.copy()
                            new_cell.alignment = cell.alignment.copy()
                            new_cell.number_format = cell.number_format

                        # 复制列宽
                        if col_idx <= source_ws.max_column:
                            ws.column_dimensions[get_column_letter(col_idx)].width = \
                                source_ws.column_dimensions[get_column_letter(col_idx)].width

                # 复制行高
                for row_idx in range(1, 5):
                    if row_idx in source_ws.row_dimensions:
                        ws.row_dimensions[row_idx].height = source_ws.row_dimensions[row_idx].height

                # 写入新数据（从第5行开始）
                for row_idx, row in enumerate(self.generated_data, start=5):
                    for col_idx, val in enumerate(row, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val if val else None)

                        # 如果原文件有对应行的样式，复制过来
                        source_row = source_ws[row_idx] if row_idx < len(source_ws) else None
                        if source_row and col_idx <= len(source_row):
                            source_cell = source_row[col_idx - 1]
                            if source_cell.has_style:
                                cell.font = source_cell.font.copy()
                                cell.fill = source_cell.fill.copy()
                                cell.border = source_cell.border.copy()
                                cell.alignment = source_cell.alignment.copy()

            else:
                # 没有原文件，创建新文件
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Wire Chart"

                # 写入表头
                header_row = ['ACTION', 'CHANGE DATE', 'CIRCUIT NBR', 'WIRE SIZE',
                             'Luxshare CABLE PART NBR', 'CUSTOMER CABLE PART NBR', 'COLOR',
                             'MARKING COLOR', 'CABLE DESG', 'EST WIRE LENGTH', 'GROUP NAME',
                             'POS NBR 1', 'CAV 1'] + self.part_numbers

                for col_idx, val in enumerate(header_row, start=1):
                    cell = ws.cell(row=4, column=col_idx, value=val)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="2170b8", end_color="2170b8", fill_type="solid")
                    cell.font = Font(color="FFFFFF")

                # 写入数据
                for row_idx, row in enumerate(self.generated_data, start=5):
                    for col_idx, val in enumerate(row, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=val if val else None)

                # 设置列宽
                for col_idx in range(1, len(header_row) + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 12

            wb.save(filepath)
            self.log(f"导出完成: {filepath}")
            messagebox.showinfo("完成", "导出成功！")

        except Exception as e:
            self.log(f"导出出错: {e}")
            messagebox.showerror("错误", f"导出失败: {e}")

    def on_mousewheel(self, event):
        """鼠标滚轮事件"""
        self.tree.yview_scroll(int(-1 * (event.delta / 120)), "units")


def main():
    root = tk.Tk()
    app = WireChartToolV5(root)
    root.mainloop()


if __name__ == "__main__":
    main()
