#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
回路表比对工具
分析新老两份回路表，根据第2列回路号进行整行比较，在新文件第一列填写变更点
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import os

def load_excel(filepath):
    """加载Excel文件"""
    return openpyxl.load_workbook(filepath)

def get_data_structure(sheet):
    """分析数据结构"""
    print(f"\n工作表名称: {sheet.title}")
    print(f"总行数: {sheet.max_row}")
    print(f"总列数: {sheet.max_column}")
    
    # 显示前5行数据
    print("\n前5行数据预览:")
    for row_idx in range(1, min(6, sheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(sheet.max_column + 1, 10)):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_data.append(str(cell_value)[:30])  # 限制每个字段最多30字符
            else:
                row_data.append("")
        print(f"第{row_idx}行: {' | '.join(row_data)}")
    
    return sheet.max_row, sheet.max_column

def compare_circuits(old_sheet, new_sheet):
    """
    比较两份回路表
    根据第2列回路号进行整行比较
    在NEW.xlsx第一列填写变更点
    """
    print("\n" + "="*80)
    print("开始比对两份回路表")
    print("="*80)
    
    # 从第4行开始读取数据（前3行为表头）
    start_row = 4
    
    # 构建旧文件的回路号字典
    old_data = {}
    old_circuit_numbers = []
    
    print(f"\n读取旧文件（从第{start_row}行开始）...")
    for row_idx in range(start_row, old_sheet.max_row + 1):
        circuit_number = old_sheet.cell(row=row_idx, column=2).value  # 第2列是回路号
        if circuit_number:
            circuit_number = str(circuit_number).strip()
            if circuit_number in old_data:
                print(f"⚠️ 警告: 旧文件中发现重复的回路号 '{circuit_number}' 在第{row_idx}行")
            old_data[circuit_number] = row_idx
            old_circuit_numbers.append(circuit_number)
    
    print(f"旧文件共读取 {len(old_data)} 条回路记录")
    
    # 构建新文件的回路号字典
    new_data = {}
    new_circuit_numbers = []
    
    print(f"\n读取新文件（从第{start_row}行开始）...")
    for row_idx in range(start_row, new_sheet.max_row + 1):
        circuit_number = new_sheet.cell(row=row_idx, column=2).value  # 第2列是回路号
        if circuit_number:
            circuit_number = str(circuit_number).strip()
            if circuit_number in new_data:
                print(f"⚠️ 警告: 新文件中发现重复的回路号 '{circuit_number}' 在第{row_idx}行")
            new_data[circuit_number] = row_idx
            new_circuit_numbers.append(circuit_number)
    
    print(f"新文件共读取 {len(new_data)} 条回路记录")
    
    # 执行比较
    added_count = 0
    deleted_count = 0
    modified_count = 0
    unchanged_count = 0
    
    print("\n" + "="*80)
    print("比对结果:")
    print("="*80)
    
    # 在新文件第一列添加表头（如果第一列是空的）
    first_header = new_sheet.cell(row=1, column=1).value
    if first_header is None:
        new_sheet.cell(row=1, column=1, value="变更点")
        new_sheet.cell(row=1, column=1).font = Font(bold=True)
    
    # 遍历新文件的每条记录
    for circuit_number in new_circuit_numbers:
        new_row_idx = new_data[circuit_number]
        old_row_idx = old_data.get(circuit_number)
        
        if old_row_idx is None:
            # 新增的回路
            added_count += 1
            new_sheet.cell(row=new_row_idx, column=1, value="✓ 新增")
            new_sheet.cell(row=new_row_idx, column=1).fill = PatternFill(
                start_color="00FF00", end_color="00FF00", fill_type="solid"
            )
            print(f"✓ 新增: 回路号 '{circuit_number}' (新文件第{new_row_idx}行)")
        else:
            # 比较整行内容
            row_changes = []
            max_col = max(old_sheet.max_column, new_sheet.max_column)
            
            for col_idx in range(1, max_col + 1):
                old_value = old_sheet.cell(row=old_row_idx, column=col_idx).value
                new_value = new_sheet.cell(row=new_row_idx, column=col_idx).value
                
                # 处理None值
                old_str = str(old_value).strip() if old_value is not None else ""
                new_str = str(new_value).strip() if new_value is not None else ""
                
                if old_str != new_str:
                    # 忽略第1列（因为我们正在写入变更标记）和空值差异
                    if col_idx != 1 and (old_str or new_str):
                        row_changes.append(f"列{col_idx}:{old_str}→{new_str}")
            
            if row_changes:
                # 变更的回路
                modified_count += 1
                change_text = "; ".join(row_changes)
                new_sheet.cell(row=new_row_idx, column=1, value=f"★ 变更: {change_text}")
                new_sheet.cell(row=new_row_idx, column=1).fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
                print(f"★ 变更: 回路号 '{circuit_number}' (新文件第{new_row_idx}行)")
                print(f"  变更内容: {change_text}")
            else:
                # 未变更
                unchanged_count += 1
                new_sheet.cell(row=new_row_idx, column=1, value="")
    
    # 标记删除的回路
    for circuit_number in old_circuit_numbers:
        if circuit_number not in new_data:
            deleted_count += 1
            print(f"✗ 删除: 回路号 '{circuit_number}' (旧文件第{old_data[circuit_number]}行)")
    
    # 打印统计信息
    print("\n" + "="*80)
    print("比对统计:")
    print("="*80)
    print(f"新增: {added_count} 条")
    print(f"删除: {deleted_count} 条")
    print(f"变更: {modified_count} 条")
    print(f"未变更: {unchanged_count} 条")
    print(f"总计: {len(new_data)} 条（新文件）")
    
    return {
        'added': added_count,
        'deleted': deleted_count,
        'modified': modified_count,
        'unchanged': unchanged_count
    }

def main():
    # 设置输出到文件
    import sys
    output_file = "比对_result.txt"
    sys.stdout = open(output_file, 'w', encoding='utf-8')
    
    print("="*80)
    print("回路表比对工具")
    print("="*80)
    print(f"当前时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 文件路径
    old_file = "比对/OLD.xlsx"
    new_file = "比对/NEW.xlsx"
    
    # 检查文件是否存在
    if not os.path.exists(old_file):
        print(f"\n❌ 错误: 找不到旧文件 '{old_file}'")
        return
    
    if not os.path.exists(new_file):
        print(f"\n❌ 错误: 找不到新文件 '{new_file}'")
        return
    
    try:
        # 加载Excel文件
        print(f"\n正在加载旧文件: {old_file}")
        old_wb = load_excel(old_file)
        old_sheet = old_wb.active
        
        print(f"\n正在加载新文件: {new_file}")
        new_wb = load_excel(new_file)
        new_sheet = new_wb.active
        
        # 分析数据结构
        print("\n" + "="*80)
        print("旧文件数据结构分析:")
        print("="*80)
        get_data_structure(old_sheet)
        
        print("\n" + "="*80)
        print("新文件数据结构分析:")
        print("="*80)
        get_data_structure(new_sheet)
        
        # 执行比较
        stats = compare_circuits(old_sheet, new_sheet)
        
        # 保存结果
        output_file = "比对/NEW_已标注.xlsx"
        new_wb.save(output_file)
        print(f"\n✓ 比对结果已保存到: {output_file}")
        
        # 关闭工作簿
        old_wb.close()
        new_wb.close()
        
        print("\n" + "="*80)
        print("比对完成！")
        print("="*80)
        
    except Exception as e:
        print(f"\n❌ 发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        # 关闭文件输出
        sys.stdout.close()
        # 恢复标准输出
        sys.stdout = sys.__stdout__
        print(f"\n比对结果已保存到: {output_file}")

if __name__ == "__main__":
    main()
