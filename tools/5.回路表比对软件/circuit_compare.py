#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
回路表比对工具 - 最终版本
分析新老两份回路表，根据第2列回路号进行整行比较，在新文件第一列填写变更点
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import os

def main():
    print("回路表比对工具")
    print("=" * 60)
    
    # 文件路径
    old_file = "比对/OLD.xlsx"
    new_file = "比对/NEW.xlsx"
    output_file = "比对/NEW_已标注.xlsx"
    
    # 检查文件
    if not os.path.exists(old_file):
        print(f"错误: 找不到 {old_file}")
        return
    if not os.path.exists(new_file):
        print(f"错误: 找不到 {new_file}")
        return
    
    print(f"加载旧文件: {old_file}")
    old_wb = openpyxl.load_workbook(old_file)
    old_sheet = old_wb.active
    print(f"  旧文件: {old_sheet.max_row} 行 x {old_sheet.max_column} 列")
    
    print(f"加载新文件: {new_file}")
    new_wb = openpyxl.load_workbook(new_file)
    new_sheet = new_wb.active
    print(f"  新文件: {new_sheet.max_row} 行 x {new_sheet.max_column} 列")
    
    # 从第4行开始读取数据
    start_row = 4
    
    # 构建旧文件数据字典
    old_data = {}
    for row_idx in range(start_row, old_sheet.max_row + 1):
        circuit_num = old_sheet.cell(row=row_idx, column=2).value
        if circuit_num:
            old_data[str(circuit_num).strip()] = row_idx
    
    print(f"旧文件读取了 {len(old_data)} 条记录")
    
    # 构建新文件数据字典并比较
    new_data = {}
    added = 0
    modified = 0
    unchanged = 0
    
    for row_idx in range(start_row, new_sheet.max_row + 1):
        circuit_num = new_sheet.cell(row=row_idx, column=2).value
        if circuit_num:
            circuit_num = str(circuit_num).strip()
            new_data[circuit_num] = row_idx
            
            if circuit_num not in old_data:
                # 新增
                added += 1
                new_sheet.cell(row=row_idx, column=1, value="✓ 新增")
                new_sheet.cell(row=row_idx, column=1).fill = PatternFill(
                    start_color="00FF00", end_color="00FF00", fill_type="solid"
                )
            else:
                # 比较整行
                old_row = old_data[circuit_num]
                changes = []
                
                for col in range(1, max(old_sheet.max_column, new_sheet.max_column) + 1):
                    old_val = old_sheet.cell(row=old_row, column=col).value
                    new_val = new_sheet.cell(row=row_idx, column=col).value
                    
                    old_str = str(old_val).strip() if old_val else ""
                    new_str = str(new_val).strip() if new_val else ""
                    
                    if col != 1 and old_str != new_str and (old_str or new_str):
                        changes.append(f"列{col}:{old_str}→{new_str}")
                
                if changes:
                    modified += 1
                    change_text = "; ".join(changes)
                    new_sheet.cell(row=row_idx, column=1, value=f"★ 变更: {change_text}")
                    new_sheet.cell(row=row_idx, column=1).fill = PatternFill(
                        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                    )
                else:
                    unchanged += 1
    
    # 计算删除
    deleted = len(old_data) - len(new_data) + added
    
    # 添加表头
    new_sheet.cell(row=1, column=1, value="变更点")
    new_sheet.cell(row=1, column=1).font = Font(bold=True)
    
    # 保存结果
    new_wb.save(output_file)
    
    # 关闭工作簿
    old_wb.close()
    new_wb.close()
    
    # 输出统计
    print("\n比对结果:")
    print("=" * 60)
    print(f"新增: {added} 条")
    print(f"删除: {deleted} 条")
    print(f"变更: {modified} 条")
    print(f"未变更: {unchanged} 条")
    print(f"总计: {len(new_data)} 条")
    print(f"\n结果已保存到: {output_file}")
    print("=" * 60)
    print("比对完成！")

if __name__ == "__main__":
    main()
