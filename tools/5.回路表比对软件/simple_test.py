#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""简单测试脚本"""

print("开始测试...")

# 测试openpyxl
try:
    import openpyxl
    print("✓ openpyxl已安装")
except ImportError:
    print("❌ openpyxl未安装")
    exit(1)

# 测试文件读取
import os

old_file = "比对/OLD.xlsx"
new_file = "比对/NEW.xlsx"

if not os.path.exists(old_file):
    print(f"❌ 找不到文件: {old_file}")
    exit(1)

if not os.path.exists(new_file):
    print(f"❌ 找不到文件: {new_file}")
    exit(1)

print(f"✓ 文件存在: {old_file}")
print(f"✓ 文件存在: {new_file}")

# 尝试读取文件
try:
    old_wb = openpyxl.load_workbook(old_file)
    old_sheet = old_wb.active
    print(f"✓ 旧文件读取成功: {old_sheet.max_row}行 x {old_sheet.max_column}列")
    
    new_wb = openpyxl.load_workbook(new_file)
    new_sheet = new_wb.active
    print(f"✓ 新文件读取成功: {new_sheet.max_row}行 x {new_sheet.max_column}列")
    
    # 显示前3行第2列的值（回路号）
    print("\n旧文件前3行的回路号（第2列）:")
    for row in range(1, min(4, old_sheet.max_row + 1)):
        value = old_sheet.cell(row=row, column=2).value
        print(f"  行{row}: {value}")
    
    print("\n新文件前3行的回路号（第2列）:")
    for row in range(1, min(4, new_sheet.max_row + 1)):
        value = new_sheet.cell(row=row, column=2).value
        print(f"  行{row}: {value}")
    
    old_wb.close()
    new_wb.close()
    
    print("\n✓ 测试完成")
    
except Exception as e:
    print(f"❌ 发生错误: {e}")
    import traceback
    traceback.print_exc()
