#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""验证Excel文件内容"""

import openpyxl
import os

print("验证Excel文件...")
print("=" * 60)

# 检查文件
old_file = "比对/OLD.xlsx"
new_file = "比对/NEW.xlsx"

print(f"\n检查文件存在性:")
print(f"OLD.xlsx: {'存在' if os.path.exists(old_file) else '不存在'}")
print(f"NEW.xlsx: {'存在' if os.path.exists(new_file) else '不存在'}")

# 读取文件
print(f"\n读取旧文件...")
try:
    old_wb = openpyxl.load_workbook(old_file)
    old_sheet = old_wb.active
    print(f"✓ 旧文件: {old_sheet.max_row} 行 x {old_sheet.max_column} 列")
    
    # 显示前5行第2列（回路号）
    print("\n旧文件前5行的第2列（回路号）:")
    for i in range(1, min(6, old_sheet.max_row + 1)):
        val = old_sheet.cell(row=i, column=2).value
        print(f"  行{i}: {val}")
    
    old_wb.close()
except Exception as e:
    print(f"✗ 错误: {e}")

print(f"\n读取新文件...")
try:
    new_wb = openpyxl.load_workbook(new_file)
    new_sheet = new_wb.active
    print(f"✓ 新文件: {new_sheet.max_row} 行 x {new_sheet.max_column} 列")
    
    # 显示前5行第2列（回路号）
    print("\n新文件前5行的第2列（回路号）:")
    for i in range(1, min(6, new_sheet.max_row + 1)):
        val = new_sheet.cell(row=i, column=2).value
        print(f"  行{i}: {val}")
    
    new_wb.close()
except Exception as e:
    print(f"✗ 错误: {e}")

print("\n" + "=" * 60)
print("验证完成！")
