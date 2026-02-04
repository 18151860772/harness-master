#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""简单检查"""
print("开始检查...")
print("Step 1: 导入openpyxl")
try:
    import openpyxl
    print("✓ openpyxl导入成功")
    print(f"  版本: {openpyxl.__version__}")
except ImportError as e:
    print(f"✗ openpyxl导入失败: {e}")
    exit(1)

print("\nStep 2: 检查文件")
import os
old_file = "比对/OLD.xlsx"
new_file = "比对/NEW.xlsx"

if os.path.exists(old_file):
    print(f"✓ {old_file} 存在")
else:
    print(f"✗ {old_file} 不存在")
    exit(1)

if os.path.exists(new_file):
    print(f"✓ {new_file} 存在")
else:
    print(f"✗ {new_file} 不存在")
    exit(1)

print("\nStep 3: 读取旧文件")
try:
    old_wb = openpyxl.load_workbook(old_file)
    old_sheet = old_wb.active
    print(f"✓ 旧文件读取成功")
    print(f"  行数: {old_sheet.max_row}")
    print(f"  列数: {old_sheet.max_column}")
except Exception as e:
    print(f"✗ 旧文件读取失败: {e}")
    exit(1)

print("\nStep 4: 读取新文件")
try:
    new_wb = openpyxl.load_workbook(new_file)
    new_sheet = new_wb.active
    print(f"✓ 新文件读取成功")
    print(f"  行数: {new_sheet.max_row}")
    print(f"  列数: {new_sheet.max_column}")
except Exception as e:
    print(f"✗ 新文件读取失败: {e}")
    exit(1)

print("\nStep 5: 保存测试文件")
try:
    test_file = "比对/NEW_已标注.xlsx"
    new_wb.save(test_file)
    print(f"✓ 测试文件保存成功: {test_file}")
except Exception as e:
    print(f"✗ 文件保存失败: {e}")
    exit(1)

print("\nStep 6: 关闭工作簿")
old_wb.close()
new_wb.close()
print("✓ 工作簿已关闭")

print("\n" + "="*60)
print("所有检查通过！")
print("="*60)
