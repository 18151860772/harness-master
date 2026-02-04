#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""测试Excel文件读取"""

try:
    import openpyxl
    print("✓ openpyxl库已成功导入")
except ImportError:
    print("❌ openpyxl库未安装")
    print("请运行: pip install openpyxl")
    exit(1)

try:
    # 读取旧文件
    print("\n读取旧文件: 比对/OLD.xlsx")
    old_wb = openpyxl.load_workbook("比对/OLD.xlsx")
    old_sheet = old_wb.active
    print(f"✓ 旧文件加载成功")
    print(f"  工作表名: {old_sheet.title}")
    print(f"  总行数: {old_sheet.max_row}")
    print(f"  总列数: {old_sheet.max_column}")
    
    # 读取新文件
    print("\n读取新文件: 比对/NEW.xlsx")
    new_wb = openpyxl.load_workbook("比对/NEW.xlsx")
    new_sheet = new_wb.active
    print(f"✓ 新文件加载成功")
    print(f"  工作表名: {new_sheet.title}")
    print(f"  总行数: {new_sheet.max_row}")
    print(f"  总列数: {new_sheet.max_column}")
    
    # 显示前5行数据
    print("\n旧文件前5行数据:")
    for row_idx in range(1, min(6, old_sheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(6, old_sheet.max_column + 1)):
            cell_value = old_sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_data.append(str(cell_value)[:20])
            else:
                row_data.append("")
        print(f"  第{row_idx}行: {' | '.join(row_data)}")
    
    print("\n新文件前5行数据:")
    for row_idx in range(1, min(6, new_sheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(6, new_sheet.max_column + 1)):
            cell_value = new_sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_data.append(str(cell_value)[:20])
            else:
                row_data.append("")
        print(f"  第{row_idx}行: {' | '.join(row_data)}")
    
    old_wb.close()
    new_wb.close()
    
    print("\n✓ 测试完成")
    
except Exception as e:
    print(f"\n❌ 发生错误: {str(e)}")
    import traceback
    traceback.print_exc()
