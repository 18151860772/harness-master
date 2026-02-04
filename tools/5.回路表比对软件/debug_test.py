#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""调试测试脚本"""

# 创建日志文件
log_file = open("debug_output.txt", 'w', encoding='utf-8')

try:
    log_file.write("开始调试测试...\n")
    
    # 测试openpyxl
    import openpyxl
    log_file.write("✓ openpyxl导入成功\n")
    
    # 测试文件读取
    import os
    old_file = "比对/OLD.xlsx"
    new_file = "比对/NEW.xlsx"
    
    log_file.write(f"检查文件: {old_file}\n")
    if os.path.exists(old_file):
        log_file.write(f"✓ {old_file} 存在\n")
    else:
        log_file.write(f"✗ {old_file} 不存在\n")
    
    log_file.write(f"检查文件: {new_file}\n")
    if os.path.exists(new_file):
        log_file.write(f"✓ {new_file} 存在\n")
    else:
        log_file.write(f"✗ {new_file} 不存在\n")
    
    # 尝试读取文件
    log_file.write("\n尝试读取旧文件...\n")
    old_wb = openpyxl.load_workbook(old_file)
    old_sheet = old_wb.active
    log_file.write(f"✓ 旧文件读取成功\n")
    log_file.write(f"  行数: {old_sheet.max_row}\n")
    log_file.write(f"  列数: {old_sheet.max_column}\n")
    
    log_file.write("\n尝试读取新文件...\n")
    new_wb = openpyxl.load_workbook(new_file)
    new_sheet = new_wb.active
    log_file.write(f"✓ 新文件读取成功\n")
    log_file.write(f"  行数: {new_sheet.max_row}\n")
    log_file.write(f"  列数: {new_sheet.max_column}\n")
    
    # 保存测试文件
    output_file = "比对/NEW_已标注.xlsx"
    new_wb.save(output_file)
    log_file.write(f"\n✓ 测试文件已保存: {output_file}\n")
    
    old_wb.close()
    new_wb.close()
    
    log_file.write("\n✓ 调试测试完成\n")
    
except Exception as e:
    log_file.write(f"\n✗ 发生错误: {str(e)}\n")
    import traceback
    log_file.write(traceback.format_exc())
finally:
    log_file.close()

print("调试测试完成，请查看 debug_output.txt")
