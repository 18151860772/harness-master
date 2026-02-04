import pandas as pd
import sys

# 设置输出编码
sys.stdout.reconfigure(encoding='utf-8')

def compare_part_numbers(master_file, diff_file, output_file):
    """
    比对零件号功能差异并生成输出文件
    
    参数:
        master_file: Master List文件路径
        diff_file: 技术差异输入文件路径
        output_file: 输出文件路径
    """
    
    print("=" * 60)
    print("零件号功能差异比对工具")
    print("=" * 60)
    
    # 读取Master List文件
    print(f"\n正在读取Master List文件: {master_file}")
    master_df = pd.read_excel(master_file, engine='openpyxl')
    print(f"Master List包含 {len(master_df)} 行数据")
    
    # 读取技术差异输入文件
    print(f"\n正在读取技术差异输入文件: {diff_file}")
    diff_df = pd.read_excel(diff_file, engine='openpyxl')
    print(f"技术差异输入包含 {len(diff_df)} 行数据")
    
    # 解析Master List数据
    master_map = {}
    for _, row in master_df.iterrows():
        part_number = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        if part_number:
            master_map[part_number] = {
                'part_number': part_number,
                'part_name': str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else '',
                'description': str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else '',
                'date': str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else '',
                'remark': str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
            }
    
    # 解析技术差异输入数据
    diff_map = {}
    for _, row in diff_df.iterrows():
        part_number = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        if part_number:
            diff_map[part_number] = {
                'part_number': part_number,
                'part_name': str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else '',
                'compare_part_number': str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else '',
                'new_features': str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else '',
                'cancelled_features': str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
            }
    
    # 比对结果
    results = []
    all_part_numbers = set(master_map.keys()) | set(diff_map.keys())
    
    stats = {
        'total': 0,
        'added': 0,
        'deleted': 0,
        'modified': 0,
        'unchanged': 0
    }
    
    print(f"\n正在比对 {len(all_part_numbers)} 个零件号...")
    
    for part_number in sorted(all_part_numbers):
        master_part = master_map.get(part_number)
        diff_part = diff_map.get(part_number)
        
        if not master_part and diff_part:
            # 新增零件
            results.append({
                '状态': '新增',
                '零件号': part_number,
                '零件名称': diff_part['part_name'],
                'Master List描述': '-',
                '对比零件号': diff_part['compare_part_number'],
                '新增功能': diff_part['new_features'],
                '取消功能': diff_part['cancelled_features'],
                '差异详情': '新增加零件'
            })
            stats['added'] += 1
            print(f"  [新增] {part_number} - {diff_part['part_name']}")
            
        elif master_part and not diff_part:
            # 删除零件
            results.append({
                '状态': '删除',
                '零件号': part_number,
                '零件名称': master_part['part_name'],
                'Master List描述': master_part['description'],
                '对比零件号': '-',
                '新增功能': '-',
                '取消功能': '-',
                '差异详情': '零件已删除'
            })
            stats['deleted'] += 1
            print(f"  [删除] {part_number} - {master_part['part_name']}")
            
        else:
            # 检查功能差异
            diff_details = []
            
            # 检查新增功能
            if diff_part['new_features'] and diff_part['new_features'] != '-':
                diff_details.append(f"新增功能: {diff_part['new_features']}")
            
            # 检查取消功能
            if diff_part['cancelled_features'] and diff_part['cancelled_features'] != '-':
                diff_details.append(f"取消功能: {diff_part['cancelled_features']}")
            
            # 检查对比零件号
            if diff_part['compare_part_number'] and diff_part['compare_part_number'] != '-':
                diff_details.append(f"对比零件号: {diff_part['compare_part_number']}")
            
            if diff_details:
                # 有功能差异
                results.append({
                    '状态': '差异',
                    '零件号': part_number,
                    '零件名称': master_part['part_name'],
                    'Master List描述': master_part['description'],
                    '对比零件号': diff_part['compare_part_number'],
                    '新增功能': diff_part['new_features'],
                    '取消功能': diff_part['cancelled_features'],
                    '差异详情': '; '.join(diff_details)
                })
                stats['modified'] += 1
                print(f"  [差异] {part_number} - {master_part['part_name']}")
            else:
                # 无差异
                results.append({
                    '状态': '无差异',
                    '零件号': part_number,
                    '零件名称': master_part['part_name'],
                    'Master List描述': master_part['description'],
                    '对比零件号': '-',
                    '新增功能': '-',
                    '取消功能': '-',
                    '差异详情': '无差异'
                })
                stats['unchanged'] += 1
        
        stats['total'] += 1
    
    # 按状态排序
    status_order = {'新增': 0, '差异': 1, '无差异': 2, '删除': 3}
    results.sort(key=lambda x: status_order.get(x['状态'], 99))
    
    # 创建DataFrame
    output_df = pd.DataFrame(results)
    
    # 保存到Excel
    print(f"\n正在保存输出文件: {output_file}")
    
    # 创建Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='零件号功能差异', index=False)
        
        # 获取工作表
        worksheet = writer.sheets['零件号功能差异']
        
        # 设置列宽
        column_widths = {
            'A': 10,  # 状态
            'B': 20,  # 零件号
            'C': 25,  # 零件名称
            'D': 40,  # Master List描述
            'E': 20,  # 对比零件号
            'F': 50,  # 新增功能
            'G': 50,  # 取消功能
            'H': 60   # 差异详情
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 添加颜色标记
        from openpyxl.styles import PatternFill, Font
        
        # 定义颜色
        fill_added = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        fill_deleted = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
        fill_modified = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        fill_unchanged = PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid")
        
        font_white = Font(color="FFFFFF", bold=True)
        
        # 设置表头样式
        for cell in worksheet[1]:
            cell.fill = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # 设置行样式
        for row_idx, row in enumerate(output_df.iterrows(), start=2):
            status = row[1]['状态']
            
            # 设置状态列颜色
            status_cell = worksheet.cell(row=row_idx, column=1)
            if status == '新增':
                status_cell.fill = fill_added
                status_cell.font = font_white
            elif status == '删除':
                status_cell.fill = fill_deleted
                status_cell.font = font_white
            elif status == '差异':
                status_cell.fill = fill_modified
                status_cell.font = font_white
            elif status == '无差异':
                status_cell.fill = fill_unchanged
                status_cell.font = font_white
    
    print(f"\n{'=' * 60}")
    print("比对完成！")
    print(f"{'=' * 60}")
    print(f"\n统计信息:")
    print(f"  总零件数: {stats['total']}")
    print(f"  新增零件: {stats['added']}")
    print(f"  删除零件: {stats['deleted']}")
    print(f"  功能差异: {stats['modified']}")
    print(f"  无差异: {stats['unchanged']}")
    print(f"\n输出文件已保存: {output_file}")
    print(f"{'=' * 60}")

if __name__ == "__main__":
    # 文件路径
    master_file = 'T28_Masterlist_20251223.xlsx'
    diff_file = '技术差异.xlsx'
    output_file = '零件号功能差异输出.xlsx'
    
    # 执行比对
    compare_part_numbers(master_file, diff_file, output_file)
