#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ECR比对工具功能测试脚本
测试ecr_compare.html工具的各项功能
"""

import os
import sys
import io
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd

# 设置标准输出编码为UTF-8
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 文件路径配置
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
HTML_FILE = os.path.join(BASE_DIR, 'ecr_compare.html')
EXCEL_FILE_1 = os.path.join(BASE_DIR, 'ECR-T28-0S_20251215.xlsx')
EXCEL_FILE_2 = os.path.join(BASE_DIR, 'ECR-T28-0S_20251230.xlsx')
REPORT_FILE = os.path.join(BASE_DIR, 'test_report.txt')

# 测试结果存储
test_results = {
    'file_existence': {},
    'html_structure': {},
    'excel_structure': {},
    'data_integrity': {},
    'data_diff': {}
}


def print_header(title):
    """打印测试标题"""
    print("\n" + "=" * 60)
    print(f"  {title}")
    print("=" * 60)


def print_test_result(test_name, passed, message=""):
    """打印测试结果"""
    status = "[PASS]" if passed else "[FAIL]"
    print(f"  {status} {test_name}")
    if message:
        print(f"    {message}")
    return passed


def check_file_existence():
    """检查文件是否存在"""
    print_header("1. 文件存在性测试")
    
    results = {}
    
    # 检查HTML文件
    html_exists = os.path.exists(HTML_FILE)
    results['html_file'] = print_test_result(
        "ecr_compare.html 存在",
        html_exists,
        f"路径: {HTML_FILE}" if html_exists else f"文件不存在: {HTML_FILE}"
    )
    
    # 检查Excel文件1
    excel1_exists = os.path.exists(EXCEL_FILE_1)
    results['excel_file_1'] = print_test_result(
        "ECR-T28-0S_20251215.xlsx 存在",
        excel1_exists,
        f"路径: {EXCEL_FILE_1}" if excel1_exists else f"文件不存在: {EXCEL_FILE_1}"
    )
    
    # 检查Excel文件2
    excel2_exists = os.path.exists(EXCEL_FILE_2)
    results['excel_file_2'] = print_test_result(
        "ECR-T28-0S_20251230.xlsx 存在",
        excel2_exists,
        f"路径: {EXCEL_FILE_2}" if excel2_exists else f"文件不存在: {EXCEL_FILE_2}"
    )
    
    test_results['file_existence'] = results
    
    # 如果文件不存在，退出测试
    if not all(results.values()):
        print("\n错误：缺少必要文件，无法继续测试")
        return False
    
    return True


def check_html_structure():
    """检查HTML文件结构"""
    print_header("2. HTML文件结构测试")
    
    results = {}
    
    try:
        with open(HTML_FILE, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # 检查SheetJS CDN引用
        sheetjs_found = 'cdn.sheetjs.com/xlsx-0.20.1/package/dist/xlsx.full.min.js' in html_content
        results['sheetjs_cdn'] = print_test_result(
            "包含SheetJS CDN引用",
            sheetjs_found,
            "xlsx-0.20.1"
        )
        
        # 检查Bootstrap 5 CDN引用
        bootstrap_css_found = 'bootstrap@5.3.0/dist/css/bootstrap.min.css' in html_content
        bootstrap_js_found = 'bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js' in html_content
        results['bootstrap_cdn'] = print_test_result(
            "包含Bootstrap 5 CDN引用",
            bootstrap_css_found and bootstrap_js_found,
            "Bootstrap 5.3.0"
        )
        
        # 检查文件上传区域
        upload_area_found = 'id="oldFileUpload"' in html_content and 'id="newFileUpload"' in html_content
        results['upload_area'] = print_test_result(
            "包含文件上传区域",
            upload_area_found,
            "旧版和新版文件上传区域"
        )
        
        # 检查比对按钮
        compare_btn_found = 'id="compareBtn"' in html_content
        results['compare_button'] = print_test_result(
            "包含比对按钮",
            compare_btn_found
        )
        
        # 检查结果展示区域
        results_section_found = 'id="resultsSection"' in html_content
        results['results_section'] = print_test_result(
            "包含结果展示区域",
            results_section_found
        )
        
        # 检查标签页
        tabs_found = 'id="diffTabs"' in html_content
        tab_types = ['新增零件', '删除零件', '修改零件']
        tabs_content = all(tab in html_content for tab in tab_types)
        results['tabs'] = print_test_result(
            "包含标签页",
            tabs_found and tabs_content,
            f"标签类型: {', '.join(tab_types)}"
        )
        
        # 检查导出按钮
        export_csv_found = 'id="exportCsvBtn"' in html_content
        export_excel_found = 'id="exportExcelBtn"' in html_content
        results['export_buttons'] = print_test_result(
            "包含导出按钮",
            export_csv_found and export_excel_found,
            "CSV和Excel导出"
        )
        
    except Exception as e:
        print(f"  ✗ 读取HTML文件失败: {str(e)}")
        return False
    
    test_results['html_structure'] = results
    return True


def check_excel_structure():
    """检查Excel文件结构"""
    print_header("3. Excel文件结构验证")
    
    results = {
        'file1': {},
        'file2': {}
    }
    
    try:
        # 检查文件1
        print("\n  文件1 (ECR-T28-0S_20251215.xlsx):")
        wb1 = load_workbook(EXCEL_FILE_1, read_only=True)
        
        # Sheet数量
        sheet_count_1 = len(wb1.sheetnames)
        results['file1']['sheet_count'] = print_test_result(
            f"Sheet数量: {sheet_count_1}",
            sheet_count_1 == 8,
            f"预期: 8, 实际: {sheet_count_1}"
        )
        
        # 第8个sheet名称
        sheet_name_1 = wb1.sheetnames[7] if len(wb1.sheetnames) > 7 else None
        results['file1']['sheet_name'] = print_test_result(
            f"第8个sheet名称: {sheet_name_1}",
            sheet_name_1 == "Masterlist-20251215",
            f"预期: Masterlist-20251215, 实际: {sheet_name_1}"
        )
        
        # 读取第8个sheet
        ws1 = wb1[sheet_name_1]
        
        # 行数
        row_count_1 = len(list(ws1.rows))
        results['file1']['row_count'] = print_test_result(
            f"行数: {row_count_1}",
            row_count_1 == 498,
            f"预期: 498, 实际: {row_count_1}"
        )
        
        # 列数
        first_row_1 = list(ws1.rows)[0]
        col_count_1 = len(first_row_1)
        results['file1']['col_count'] = print_test_result(
            f"列数: {col_count_1}",
            col_count_1 == 419,
            f"预期: 419, 实际: {col_count_1}"
        )
        
        # 验证关键列
        headers_1 = [cell.value for cell in first_row_1]
        key_columns = ['零件号', '零件名称', '用量', '功能']
        key_columns_found = all(col in headers_1 for col in key_columns)
        results['file1']['key_columns'] = print_test_result(
            "关键列存在",
            key_columns_found,
            f"检查列: {', '.join(key_columns)}"
        )
        
        wb1.close()
        
        # 检查文件2
        print("\n  文件2 (ECR-T28-0S_20251230.xlsx):")
        wb2 = load_workbook(EXCEL_FILE_2, read_only=True)
        
        # Sheet数量
        sheet_count_2 = len(wb2.sheetnames)
        results['file2']['sheet_count'] = print_test_result(
            f"Sheet数量: {sheet_count_2}",
            sheet_count_2 == 9,
            f"预期: 9, 实际: {sheet_count_2}"
        )
        
        # 第8个sheet名称
        sheet_name_2 = wb2.sheetnames[7] if len(wb2.sheetnames) > 7 else None
        results['file2']['sheet_name'] = print_test_result(
            f"第8个sheet名称: {sheet_name_2}",
            sheet_name_2 == "Masterlist-20251229",
            f"预期: Masterlist-20251229, 实际: {sheet_name_2}"
        )
        
        # 读取第8个sheet
        ws2 = wb2[sheet_name_2]
        
        # 行数
        row_count_2 = len(list(ws2.rows))
        results['file2']['row_count'] = print_test_result(
            f"行数: {row_count_2}",
            row_count_2 == 512,
            f"预期: 512, 实际: {row_count_2}"
        )
        
        # 列数
        first_row_2 = list(ws2.rows)[0]
        col_count_2 = len(first_row_2)
        results['file2']['col_count'] = print_test_result(
            f"列数: {col_count_2}",
            col_count_2 == 419,
            f"预期: 419, 实际: {col_count_2}"
        )
        
        # 验证关键列
        headers_2 = [cell.value for cell in first_row_2]
        key_columns_found_2 = all(col in headers_2 for col in key_columns)
        results['file2']['key_columns'] = print_test_result(
            "关键列存在",
            key_columns_found_2,
            f"检查列: {', '.join(key_columns)}"
        )
        
        wb2.close()
        
    except Exception as e:
        print(f"  ✗ 读取Excel文件失败: {str(e)}")
        return False
    
    test_results['excel_structure'] = results
    return True


def check_data_integrity():
    """检查数据完整性"""
    print_header("4. 数据完整性测试")
    
    results = {}
    
    try:
        # 读取两个文件的第8个sheet
        df1 = pd.read_excel(EXCEL_FILE_1, sheet_name=7, header=0)
        df2 = pd.read_excel(EXCEL_FILE_2, sheet_name=7, header=0)
        
        # 获取零件号列
        part_numbers_1 = df1.iloc[:, 0].dropna()  # 第一列为零件号
        part_numbers_2 = df2.iloc[:, 0].dropna()
        
        # 检查零件号是否为主键（无重复）
        duplicates_1 = part_numbers_1.duplicated().sum()
        results['file1_no_duplicates'] = print_test_result(
            f"文件1零件号无重复",
            duplicates_1 == 0,
            f"重复数量: {duplicates_1}"
        )
        
        duplicates_2 = part_numbers_2.duplicated().sum()
        results['file2_no_duplicates'] = print_test_result(
            f"文件2零件号无重复",
            duplicates_2 == 0,
            f"重复数量: {duplicates_2}"
        )
        
        # 统计零件数量
        part_count_1 = len(part_numbers_1)
        part_count_2 = len(part_numbers_2)
        
        print(f"\n  文件1零件数量: {part_count_1} (预期: 497)")
        print(f"  文件2零件数量: {part_count_2} (预期: 511)")
        
        results['file1_part_count'] = part_count_1
        results['file2_part_count'] = part_count_2
        
        # 保存数据用于后续差异计算
        test_results['data_integrity'] = {
            'file1_no_duplicates': duplicates_1 == 0,
            'file2_no_duplicates': duplicates_2 == 0,
            'file1_part_count': part_count_1,
            'file2_part_count': part_count_2,
            'df1': df1,
            'df2': df2,
            'part_numbers_1': part_numbers_1,
            'part_numbers_2': part_numbers_2
        }
        
    except Exception as e:
        print(f"  ✗ 数据完整性检查失败: {str(e)}")
        return False
    
    return True


def calculate_data_diff():
    """预计算数据差异"""
    print_header("5. 数据差异预计算")
    
    try:
        df1 = test_results['data_integrity']['df1']
        df2 = test_results['data_integrity']['df2']
        
        # 获取零件号列（第一列）
        part_numbers_1 = set(df1.iloc[:, 0].dropna().astype(str))
        part_numbers_2 = set(df2.iloc[:, 0].dropna().astype(str))
        
        # 识别新增零件（文件2有，文件1没有）
        added_parts = part_numbers_2 - part_numbers_1
        
        # 识别删除零件（文件1有，文件2没有）
        deleted_parts = part_numbers_1 - part_numbers_2
        
        # 识别共同零件
        common_parts = part_numbers_1 & part_numbers_2
        
        # 识别修改零件
        modified_parts = set()
        unchanged_parts = set()
        
        # 设置零件号为索引
        df1_indexed = df1.set_index(df1.columns[0])
        df2_indexed = df2.set_index(df2.columns[0])
        
        # 比较共同零件
        for part in common_parts:
            if part in df1_indexed.index and part in df2_indexed.index:
                row1 = df1_indexed.loc[part]
                row2 = df2_indexed.loc[part]
                
                # 比较所有列
                is_modified = False
                for col in df1.columns:
                    val1 = row1[col] if col in row1.index else None
                    val2 = row2[col] if col in row2.index else None
                    
                    # 标准化值进行比较
                    val1_norm = str(val1).strip() if pd.notna(val1) else ''
                    val2_norm = str(val2).strip() if pd.notna(val2) else ''
                    
                    if val1_norm != val2_norm:
                        is_modified = True
                        break
                
                if is_modified:
                    modified_parts.add(part)
                else:
                    unchanged_parts.add(part)
        
        # 统计结果
        added_count = len(added_parts)
        deleted_count = len(deleted_parts)
        modified_count = len(modified_parts)
        unchanged_count = len(unchanged_parts)
        
        print(f"  新增零件: {added_count}个")
        print(f"  删除零件: {deleted_count}个")
        print(f"  修改零件: {modified_count}个")
        print(f"  无变化零件: {unchanged_count}个")
        
        # 保存差异结果
        test_results['data_diff'] = {
            'added_parts': added_parts,
            'deleted_parts': deleted_parts,
            'modified_parts': modified_parts,
            'unchanged_parts': unchanged_parts,
            'added_count': added_count,
            'deleted_count': deleted_count,
            'modified_count': modified_count,
            'unchanged_count': unchanged_count
        }
        
        # 显示部分新增零件示例
        if added_parts:
            print(f"\n  新增零件示例（前5个）:")
            for i, part in enumerate(list(added_parts)[:5], 1):
                print(f"    {i}. {part}")
        
        # 显示部分删除零件示例
        if deleted_parts:
            print(f"\n  删除零件示例（前5个）:")
            for i, part in enumerate(list(deleted_parts)[:5], 1):
                print(f"    {i}. {part}")
        
        # 显示部分修改零件示例
        if modified_parts:
            print(f"\n  修改零件示例（前5个）:")
            for i, part in enumerate(list(modified_parts)[:5], 1):
                print(f"    {i}. {part}")
        
        return True
        
    except Exception as e:
        print(f"  ✗ 数据差异计算失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def generate_report():
    """生成测试报告"""
    print_header("6. 生成测试报告")
    
    try:
        report_lines = []
        
        # 报告标题
        report_lines.append("=" * 60)
        report_lines.append("ECR比对工具功能测试报告")
        report_lines.append("=" * 60)
        report_lines.append("")
        
        # 测试时间
        test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        report_lines.append(f"测试时间：{test_time}")
        report_lines.append("")
        
        # 1. 文件存在性测试
        report_lines.append("1. 文件存在性测试")
        report_lines.append("-" * 40)
        file_results = test_results['file_existence']
        report_lines.append(f"  {'[PASS]' if file_results.get('html_file', False) else '[FAIL]'} ecr_compare.html 存在")
        report_lines.append(f"  {'[PASS]' if file_results.get('excel_file_1', False) else '[FAIL]'} ECR-T28-0S_20251215.xlsx 存在")
        report_lines.append(f"  {'[PASS]' if file_results.get('excel_file_2', False) else '[FAIL]'} ECR-T28-0S_20251230.xlsx 存在")
        report_lines.append("")
        
        # 2. HTML文件结构测试
        report_lines.append("2. HTML文件结构测试")
        report_lines.append("-" * 40)
        html_results = test_results['html_structure']
        report_lines.append(f"  {'[PASS]' if html_results.get('sheetjs_cdn', False) else '[FAIL]'} 包含SheetJS CDN引用")
        report_lines.append(f"  {'[PASS]' if html_results.get('bootstrap_cdn', False) else '[FAIL]'} 包含Bootstrap 5 CDN引用")
        report_lines.append(f"  {'[PASS]' if html_results.get('upload_area', False) else '[FAIL]'} 包含文件上传区域")
        report_lines.append(f"  {'[PASS]' if html_results.get('compare_button', False) else '[FAIL]'} 包含比对按钮")
        report_lines.append(f"  {'[PASS]' if html_results.get('results_section', False) else '[FAIL]'} 包含结果展示区域")
        report_lines.append(f"  {'[PASS]' if html_results.get('tabs', False) else '[FAIL]'} 包含标签页")
        report_lines.append(f"  {'[PASS]' if html_results.get('export_buttons', False) else '[FAIL]'} 包含导出按钮")
        report_lines.append("")
        
        # 3. Excel文件结构验证
        report_lines.append("3. Excel文件结构验证")
        report_lines.append("-" * 40)
        
        excel_results = test_results['excel_structure']
        
        # 文件1
        report_lines.append("  文件1 (ECR-T28-0S_20251215.xlsx):")
        report_lines.append(f"    {'[PASS]' if excel_results['file1'].get('sheet_count', False) else '[FAIL]'} Sheet数量: 8")
        report_lines.append(f"    {'[PASS]' if excel_results['file1'].get('sheet_name', False) else '[FAIL]'} 第8个sheet名称: Masterlist-20251215")
        report_lines.append(f"    {'[PASS]' if excel_results['file1'].get('row_count', False) else '[FAIL]'} 行数: 498")
        report_lines.append(f"    {'[PASS]' if excel_results['file1'].get('col_count', False) else '[FAIL]'} 列数: 419")
        report_lines.append(f"    {'[PASS]' if excel_results['file1'].get('key_columns', False) else '[FAIL]'} 关键列存在")
        report_lines.append("")
        
        # 文件2
        report_lines.append("  文件2 (ECR-T28-0S_20251230.xlsx):")
        report_lines.append(f"    {'[PASS]' if excel_results['file2'].get('sheet_count', False) else '[FAIL]'} Sheet数量: 9")
        report_lines.append(f"    {'[PASS]' if excel_results['file2'].get('sheet_name', False) else '[FAIL]'} 第8个sheet名称: Masterlist-20251229")
        report_lines.append(f"    {'[PASS]' if excel_results['file2'].get('row_count', False) else '[FAIL]'} 行数: 512")
        report_lines.append(f"    {'[PASS]' if excel_results['file2'].get('col_count', False) else '[FAIL]'} 列数: 419")
        report_lines.append(f"    {'[PASS]' if excel_results['file2'].get('key_columns', False) else '[FAIL]'} 关键列存在")
        report_lines.append("")
        
        # 4. 数据完整性测试
        report_lines.append("4. 数据完整性测试")
        report_lines.append("-" * 40)
        data_results = test_results['data_integrity']
        report_lines.append(f"  {'[PASS]' if data_results.get('file1_no_duplicates', False) else '[FAIL]'} 文件1零件号无重复")
        report_lines.append(f"  {'[PASS]' if data_results.get('file2_no_duplicates', False) else '[FAIL]'} 文件2零件号无重复")
        report_lines.append(f"  文件1零件数量: {data_results.get('file1_part_count', 0)}")
        report_lines.append(f"  文件2零件数量: {data_results.get('file2_part_count', 0)}")
        report_lines.append("")
        
        # 5. 数据差异预计算
        report_lines.append("5. 数据差异预计算")
        report_lines.append("-" * 40)
        diff_results = test_results['data_diff']
        report_lines.append(f"  新增零件: {diff_results.get('added_count', 0)}个")
        report_lines.append(f"  删除零件: {diff_results.get('deleted_count', 0)}个")
        report_lines.append(f"  修改零件: {diff_results.get('modified_count', 0)}个")
        report_lines.append(f"  无变化零件: {diff_results.get('unchanged_count', 0)}个")
        report_lines.append("")
        
        # 新增零件详情
        if diff_results.get('added_parts'):
            report_lines.append("  新增零件列表:")
            for i, part in enumerate(sorted(diff_results['added_parts']), 1):
                report_lines.append(f"    {i}. {part}")
            report_lines.append("")
        
        # 删除零件详情
        if diff_results.get('deleted_parts'):
            report_lines.append("  删除零件列表:")
            for i, part in enumerate(sorted(diff_results['deleted_parts']), 1):
                report_lines.append(f"    {i}. {part}")
            report_lines.append("")
        
        # 修改零件详情
        if diff_results.get('modified_parts'):
            report_lines.append("  修改零件列表:")
            for i, part in enumerate(sorted(diff_results['modified_parts']), 1):
                report_lines.append(f"    {i}. {part}")
            report_lines.append("")
        
        # 测试结论
        report_lines.append("=" * 60)
        report_lines.append("测试结论")
        report_lines.append("=" * 60)
        
        # 计算通过率
        all_tests = []
        all_tests.extend(file_results.values())
        all_tests.extend(html_results.values())
        all_tests.extend(excel_results['file1'].values())
        all_tests.extend(excel_results['file2'].values())
        all_tests.append(data_results.get('file1_no_duplicates', False))
        all_tests.append(data_results.get('file2_no_duplicates', False))
        
        passed_tests = sum(all_tests)
        total_tests = len(all_tests)
        pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        report_lines.append(f"总测试项: {total_tests}")
        report_lines.append(f"通过: {passed_tests}")
        report_lines.append(f"失败: {total_tests - passed_tests}")
        report_lines.append(f"通过率: {pass_rate:.2f}%")
        report_lines.append("")
        
        if pass_rate == 100:
            report_lines.append("测试结论：所有测试通过 [OK]")
        elif pass_rate >= 80:
            report_lines.append("测试结论：大部分测试通过，存在少量问题 [WARN]")
        else:
            report_lines.append("测试结论：存在较多问题，需要修复 [FAIL]")
        
        # 建议
        report_lines.append("")
        report_lines.append("=" * 60)
        report_lines.append("建议")
        report_lines.append("=" * 60)
        
        if pass_rate == 100:
            report_lines.append("- 工具功能完整，可以进行正常使用")
            report_lines.append("- 建议定期运行测试脚本验证功能")
        else:
            report_lines.append("- 请检查失败的测试项")
            report_lines.append("- 根据测试结果修复相应问题")
            report_lines.append("- 修复后重新运行测试脚本")
        
        # 写入报告文件
        with open(REPORT_FILE, 'w', encoding='utf-8') as f:
            f.write('\n'.join(report_lines))
        
        print(f"  ✓ 测试报告已生成: {REPORT_FILE}")
        print(f"\n报告内容预览:")
        print('\n'.join(report_lines[:50]))  # 显示前50行
        
        return True
        
    except Exception as e:
        print(f"  ✗ 生成测试报告失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """主函数"""
    print("\n" + "=" * 60)
    print("  ECR比对工具功能测试")
    print("=" * 60)
    
    # 1. 检查文件存在性
    if not check_file_existence():
        print("\n测试终止：缺少必要文件")
        return False
    
    # 2. 检查HTML结构
    if not check_html_structure():
        print("\n测试终止：HTML结构检查失败")
        return False
    
    # 3. 检查Excel结构
    if not check_excel_structure():
        print("\n测试终止：Excel结构检查失败")
        return False
    
    # 4. 检查数据完整性
    if not check_data_integrity():
        print("\n测试终止：数据完整性检查失败")
        return False
    
    # 5. 计算数据差异
    if not calculate_data_diff():
        print("\n测试终止：数据差异计算失败")
        return False
    
    # 6. 生成测试报告
    if not generate_report():
        print("\n测试终止：生成测试报告失败")
        return False
    
    print("\n" + "=" * 60)
    print("  测试完成！")
    print("=" * 60)
    
    return True


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
