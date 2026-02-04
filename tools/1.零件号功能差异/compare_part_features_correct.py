import pandas as pd
import sys

# 设置输出编码
sys.stdout.reconfigure(encoding='utf-8')

def compare_part_features_correct(master_file, diff_file, output_file):
    """
    比对零件号之间的功能差异（基于●标记）
    输出格式与技术差异输入文件一致（5列）
    差异内容为功能列的值（从第二行获取）
    
    参数:
        master_file: Master List文件路径
        diff_file: 技术差异输入文件路径
        output_file: 输出文件路径
    """
    
    print("=" * 80)
    print("零件号功能差异比对工具（基于●标记）")
    print("=" * 80)
    
    # 读取Master List文件
    print(f"\n正在读取Master List文件: {master_file}")
    master_df = pd.read_excel(master_file, engine='openpyxl')
    print(f"Master List包含 {len(master_df)} 行数据，{len(master_df.columns)} 列")
    
    # 读取技术差异输入文件
    print(f"\n正在读取技术差异输入文件: {diff_file}")
    diff_df = pd.read_excel(diff_file, engine='openpyxl')
    print(f"技术差异输入包含 {len(diff_df)} 行数据")
    
    # 获取第二行数据（索引1），第一行是表头
    second_row = master_df.iloc[1]
    
    # 查找"标配"单元格的位置
    standard_column_index = None
    for i, value in enumerate(second_row):
        if pd.notna(value) and '标配' in str(value):
            standard_column_index = i
            print(f"\n找到'标配'单元格在第二行索引 {standard_column_index}: {value}")
            break
    
    if standard_column_index is None:
        print("\n警告：未在第二行找到'标配'单元格")
        return
    
    # 获取功能列（从"标配"单元格之后开始）
    feature_columns = master_df.columns[standard_column_index + 1:].tolist()
    print(f"从'标配'单元格之后（索引 {standard_column_index + 1}）识别到 {len(feature_columns)} 个功能/配置列")
    
    # 解析Master List数据，建立零件号到功能的映射
    print("\n正在解析Master List数据...")
    part_map = {}
    for _, row in master_df.iterrows():
        part_number = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        if part_number and part_number != 'nan':
            features = set()
            for col in feature_columns:
                value = row[col]
                # 检查是否包含●标记
                if pd.notna(value) and '●' in str(value):
                    features.add(col)
            part_map[part_number] = {
                'part_number': part_number,
                'part_name': str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else '',
                'status': str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else '',
                'quantity': str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else '',
                'description': str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else '',
                'features': features
            }
    
    print(f"解析完成，共识别 {len(part_map)} 个零件号")
    
    # 处理技术差异输入文件中的每个比对项
    print("\n正在比对零件号功能差异...")
    results = []
    
    for _, diff_row in diff_df.iterrows():
        part_number1 = str(diff_row.iloc[0]).strip() if pd.notna(diff_row.iloc[0]) else ''
        part_name1 = str(diff_row.iloc[1]).strip() if pd.notna(diff_row.iloc[1]) else ''
        part_number2 = str(diff_row.iloc[2]).strip() if pd.notna(diff_row.iloc[2]) else ''
        
        if not part_number1:
            print(f"  跳过：零件号为空")
            continue
        
        part1 = part_map.get(part_number1)
        part2 = part_map.get(part_number2) if part_number2 else None
        
        if not part1:
            print(f"  警告：零件号 {part_number1} 在Master List中未找到")
            continue
        
        print(f"\n  比对: {part_number1} ({part1['part_name']})")
        if part2:
            print(f"       vs {part_number2} ({part2['part_name']})")
        else:
            print(f"       vs {part_number2} (未找到)")
        
        # 获取两个零件的功能
        features1 = part1['features']
        features2 = part2['features'] if part2 else set()
        
        # 识别差异
        # 新增功能：零件号1有但零件号2没有的功能
        added_features = features1 - features2
        # 取消功能：零件号2有但零件号1没有的功能
        cancelled_features = features2 - features1
        
        # 将功能列表转换为字符串（使用功能列的值，用●分隔）
        added_features_str = '●'.join(sorted(added_features)) if added_features else ''
        cancelled_features_str = '●'.join(sorted(cancelled_features)) if cancelled_features else ''
        
        # 添加结果（与技术差异输入文件格式一致：5列）
        results.append({
            '零件号': part_number1,
            '零件名称(中文）': part1['part_name'],
            '对比零件号': part_number2,
            '新增功能': added_features_str,
            '取消功能': cancelled_features_str
        })
        
        print(f"    结果: 新增 {len(added_features)} 项, 取消 {len(cancelled_features)} 项")
        if added_features:
            print(f"    新增功能示例: {list(added_features)[:3]}")
        if cancelled_features:
            print(f"    取消功能示例: {list(cancelled_features)[:3]}")
    
    # 创建DataFrame（与技术差异输入文件格式一致）
    output_df = pd.DataFrame(results)
    
    # 保存到Excel
    print(f"\n正在保存输出文件: {output_file}")
    
    # 创建Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='零件号功能差异', index=False)
        
        # 获取工作表
        worksheet = writer.sheets['零件号功能差异']
        
        # 设置列宽
        column_widths = {
            'A': 20,  # 零件号
            'B': 25,  # 零件名称(中文）
            'C': 20,  # 对比零件号
            'D': 80,  # 新增功能
            'E': 80   # 取消功能
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置行高（用于显示多行文本）
        for row in range(2, len(results) + 2):
            worksheet.row_dimensions[row].height = 100
        
        # 设置表头样式
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        header_fill = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 设置数据单元格样式（自动换行）
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border
        
        # 为新增功能和取消功能列添加颜色标记
        for row_idx, result in enumerate(results, start=2):
            # 新增功能列（D列）
            if result['新增功能']:
                cell = worksheet.cell(row=row_idx, column=4)
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                cell.font = Font(color="1B5E20", bold=True)
            
            # 取消功能列（E列）
            if result['取消功能']:
                cell = worksheet.cell(row=row_idx, column=5)
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                cell.font = Font(color="B71C1C", bold=True)
    
    print(f"\n{'=' * 80}")
    print("比对完成！")
    print(f"{'=' * 80}")
    print(f"\n统计信息:")
    print(f"  比对了 {len(results)} 组零件号")
    print(f"\n输出文件已保存: {output_file}")
    print(f"{'=' * 80}")

if __name__ == "__main__":
    # 文件路径
    master_file = 'T28_Masterlist_20251223.xlsx'
    diff_file = '技术差异.xlsx'
    output_file = '零件号功能差异输出.xlsx'
    
    # 执行比对
    compare_part_features_correct(master_file, diff_file, output_file)
