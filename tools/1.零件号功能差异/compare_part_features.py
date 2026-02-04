import pandas as pd
import sys

# 设置输出编码
sys.stdout.reconfigure(encoding='utf-8')

def compare_part_features(master_file, diff_file, output_file):
    """
    比对零件号之间的功能差异（基于●标记）
    
    参数:
        master_file: Master List文件路径
        diff_file: 技术差异输入文件路径
        output_file: 输出文件路径
    """
    
    print("=" * 80)
    print("零件号功能差异比对工具（基于●标记）")
    print("=" * 80)
    
    # 读取Master List文件
    print(f"\n正在读取Master List文件: {master_file}")
    master_df = pd.read_excel(master_file, engine='openpyxl')
    print(f"Master List包含 {len(master_df)} 行数据，{len(master_df.columns)} 列")
    
    # 读取技术差异输入文件
    print(f"\n正在读取技术差异输入文件: {diff_file}")
    diff_df = pd.read_excel(diff_file, engine='openpyxl')
    print(f"技术差异输入包含 {len(diff_df)} 行数据")
    
    # 获取功能列（从"功能"列之后开始）
    # 基本信息：零件号、零件名称(中文）、沿用/新开发/预留/删除、用量、功能
    # 从"功能"列之后（索引5）开始是功能配置列
    feature_columns = master_df.columns[5:].tolist()
    print(f"\n识别到 {len(feature_columns)} 个功能/配置列")
    
    # 解析Master List数据，建立零件号到功能的映射
    print("\n正在解析Master List数据...")
    part_map = {}
    for _, row in master_df.iterrows():
        part_number = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        if part_number and part_number != 'nan':
            features = {}
            for col in feature_columns:
                value = row[col]
                # 检查是否包含●标记
                if pd.notna(value) and '●' in str(value):
                    features[col] = str(value).strip()
            part_map[part_number] = {
                'part_number': part_number,
                'part_name': str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else '',
                'status': str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else '',
                'quantity': str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else '',
                'description': str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else '',
                'features': features
            }
    
    print(f"解析完成，共识别 {len(part_map)} 个零件号")
    
    # 处理技术差异输入文件中的每个比对项
    print("\n正在比对零件号功能差异...")
    results = []
    
    for _, diff_row in diff_df.iterrows():
        part_number1 = str(diff_row.iloc[0]).strip() if pd.notna(diff_row.iloc[0]) else ''
        part_name1 = str(diff_row.iloc[1]).strip() if pd.notna(diff_row.iloc[1]) else ''
        part_number2 = str(diff_row.iloc[2]).strip() if pd.notna(diff_row.iloc[2]) else ''
        
        if not part_number1:
            print(f"  跳过：零件号为空")
            continue
        
        part1 = part_map.get(part_number1)
        part2 = part_map.get(part_number2) if part_number2 else None
        
        if not part1:
            print(f"  警告：零件号 {part_number1} 在Master List中未找到")
            continue
        
        print(f"\n  比对: {part_number1} ({part1['part_name']})")
        if part2:
            print(f"       vs {part_number2} ({part2['part_name']})")
        else:
            print(f"       vs {part_number2} (未找到)")
        
        # 获取两个零件的功能
        features1 = part1['features']
        features2 = part2['features'] if part2 else {}
        
        # 找出所有功能列
        all_feature_columns = set(features1.keys()) | set(features2.keys())
        
        # 识别差异
        only_in_part1 = []
        only_in_part2 = []
        common_features = []
        
        for feature_col in sorted(all_feature_columns):
            in_part1 = feature_col in features1
            in_part2 = feature_col in features2
            
            if in_part1 and not in_part2:
                only_in_part1.append(feature_col)
            elif not in_part1 and in_part2:
                only_in_part2.append(feature_col)
            elif in_part1 and in_part2:
                # 两个零件都有此功能
                common_features.append(feature_col)
        
        # 生成差异详情
        diff_details = []
        
        if only_in_part1:
            diff_details.append(f"仅{part_number1}有: {', '.join(only_in_part1[:5])}")
            if len(only_in_part1) > 5:
                diff_details[-1] += f" 等{len(only_in_part1)}项"
        
        if only_in_part2:
            diff_details.append(f"仅{part_number2}有: {', '.join(only_in_part2[:5])}")
            if len(only_in_part2) > 5:
                diff_details[-1] += f" 等{len(only_in_part2)}项"
        
        if not diff_details:
            diff_details.append("功能完全相同")
        
        # 添加结果
        result = {
            '零件号1': part_number1,
            '零件名称1': part1['part_name'],
            '状态1': part1['status'],
            '用量1': part1['quantity'],
            '描述1': part1['description'],
            '零件号2': part_number2,
            '零件名称2': part2['part_name'] if part2 else '-',
            '状态2': part2['status'] if part2 else '-',
            '用量2': part2['quantity'] if part2 else '-',
            '描述2': part2['description'] if part2 else '-',
            '仅零件号1有的功能': '\n'.join(only_in_part1) if only_in_part1 else '-',
            '仅零件号2有的功能': '\n'.join(only_in_part2) if only_in_part2 else '-',
            '共同功能': '\n'.join(common_features) if common_features else '-',
            '差异统计': f"零件1独有: {len(only_in_part1)}项, 零件2独有: {len(only_in_part2)}项, 共同: {len(common_features)}项",
            '差异详情': '; '.join(diff_details)
        }
        results.append(result)
        
        print(f"    结果: 零件1独有 {len(only_in_part1)} 项, 零件2独有 {len(only_in_part2)} 项, 共同 {len(common_features)} 项")
    
    # 创建DataFrame
    if results:
        output_df = pd.DataFrame(results)
        
        # 保存到Excel
        print(f"\n正在保存输出文件: {output_file}")
        
        # 创建Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            output_df.to_excel(writer, sheet_name='零件号功能差异', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['零件号功能差异']
            
            # 设置列宽
            column_widths = {
                'A': 15,  # 零件号1
                'B': 20,  # 零件名称1
                'C': 15,  # 状态1
                'D': 10,  # 用量1
                'E': 40,  # 描述1
                'F': 15,  # 零件号2
                'G': 20,  # 零件名称2
                'H': 15,  # 状态2
                'I': 10,  # 用量2
                'J': 40,  # 描述2
                'K': 60,  # 仅零件号1有的功能
                'L': 60,  # 仅零件号2有的功能
                'M': 60,  # 共同功能
                'N': 40,  # 差异统计
                'O': 80   # 差异详情
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置行高（用于显示多行文本）
            for row in range(2, len(results) + 2):
                worksheet.row_dimensions[row].height = 100
            
            # 设置表头样式
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 设置数据单元格样式（自动换行）
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = thin_border
        
        print(f"\n{'=' * 80}")
        print("比对完成！")
        print(f"{'=' * 80}")
        print(f"\n统计信息:")
        print(f"  比对了 {len(results)} 组零件号")
        print(f"\n输出文件已保存: {output_file}")
        print(f"{'=' * 80}")
    else:
        print("\n没有找到需要比对的零件号")

if __name__ == "__main__":
    # 文件路径
    master_file = 'T28_Masterlist_20251223.xlsx'
    diff_file = '技术差异.xlsx'
    output_file = '零件号功能差异输出.xlsx'
    
    # 执行比对
    compare_part_features(master_file, diff_file, output_file)
