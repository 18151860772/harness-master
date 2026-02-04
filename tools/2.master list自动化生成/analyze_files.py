import pandas as pd
import openpyxl

# 读取Wire list文件
print("=" * 80)
print("Wire List 文件结构分析")
print("=" * 80)

wire_list_file = "Wire list.xlsx"
try:
    # 使用openpyxl读取以查看所有sheet
    wb = openpyxl.load_workbook(wire_list_file, data_only=True)
    print(f"\nSheet列表: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        print(f"数据行数: {ws.max_row}, 列数: {ws.max_column}")
        
        # 显示前10行数据
        print("\n前10行数据:")
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= 10:
                print(f"Row {i}: {row}")
            else:
                break
    
    # 使用pandas读取详细数据
    for sheet_name in wb.sheetnames:
        df = pd.read_excel(wire_list_file, sheet_name=sheet_name, header=None)
        print(f"\n--- Sheet '{sheet_name}' 详细列信息 ---")
        print(f"形状: {df.shape}")
        print(f"列数: {len(df.columns)}")
        print("\n前5行数据:")
        print(df.head())
        
except Exception as e:
    print(f"读取Wire list文件时出错: {e}")

print("\n" + "=" * 80)
print("配置表 文件结构分析")
print("=" * 80)

config_file = "配置表.xlsx"
try:
    # 使用openpyxl读取以查看所有sheet
    wb = openpyxl.load_workbook(config_file, data_only=True)
    print(f"\nSheet列表: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        print(f"数据行数: {ws.max_row}, 列数: {ws.max_column}")
        
        # 显示前10行数据
        print("\n前10行数据:")
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= 10:
                print(f"Row {i}: {row}")
            else:
                break
    
    # 使用pandas读取详细数据
    for sheet_name in wb.sheetnames:
        df = pd.read_excel(config_file, sheet_name=sheet_name, header=None)
        print(f"\n--- Sheet '{sheet_name}' 详细列信息 ---")
        print(f"形状: {df.shape}")
        print(f"列数: {len(df.columns)}")
        print("\n前5行数据:")
        print(df.head())
        
except Exception as e:
    print(f"读取配置表文件时出错: {e}")
