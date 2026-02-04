import pandas as pd
import openpyxl

# 读取现有的MasterList文件
masterlist_file = "MasterList__202508060203859.xlsx"

try:
    # 使用openpyxl读取
    wb = openpyxl.load_workbook(masterlist_file, data_only=True)
    
    print("=" * 80)
    print("现有MasterList文件格式分析")
    print("=" * 80)
    
    print(f"\nSheet列表: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n{'=' * 80}")
        print(f"Sheet: {sheet_name}")
        print(f"{'=' * 80}")
        
        ws = wb[sheet_name]
        print(f"数据行数: {ws.max_row}, 列数: {ws.max_column}")
        
        # 显示前20行数据
        print("\n前20行数据:")
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= 20:
                print(f"Row {i}: {row}")
            else:
                break
    
    # 使用pandas读取详细数据
    for sheet_name in wb.sheetnames:
        df = pd.read_excel(masterlist_file, sheet_name=sheet_name, header=None)
        print(f"\n{'=' * 80}")
        print(f"Sheet '{sheet_name}' 详细分析")
        print(f"{'=' * 80}")
        print(f"形状: {df.shape}")
        print(f"列数: {len(df.columns)}")
        
        print("\n前10行数据:")
        print(df.head(10))
        
        print("\n列信息:")
        for i in range(min(20, len(df.columns))):
            print(f"列 {i}: {df.iloc[:5, i].tolist()}")
        
except Exception as e:
    print(f"出错: {e}")
    import traceback
    traceback.print_exc()
