#!/usr/bin/env python
# -*- coding: utf-8 -*-

try:
    from openpyxl import load_workbook

    # 读取Excel文件
    file_path = r'C:\Users\HP\Desktop\tools\18.保险丝线径匹配评审\inline.xlsx'
    wb = load_workbook(file_path, data_only=True)

    print("=" * 80)
    print("Excel文件分析报告")
    print("=" * 80)

    # 显示所有工作表
    print(f"\n工作表列表: {wb.sheetnames}")

    # 读取第一个工作表
    ws = wb.active
    print(f"\n当前工作表: {ws.title}")
    print(f"数据范围: {ws.max_row} 行 x {ws.max_column} 列")

    # 读取前20行数据
    print("\n" + "=" * 80)
    print("数据预览（前20行）:")
    print("=" * 80)

    data = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        data.append(row)
        if row_idx <= 20:
            print(f"\n第{row_idx}行:")
            for col_idx, cell_value in enumerate(row, 1):
                print(f"  列{col_idx}: {cell_value}")

    # 分析列名
    print("\n" + "=" * 80)
    print("列名分析:")
    print("=" * 80)

    if len(data) > 0:
        headers = data[0]
        print(f"\n总共有 {len(headers)} 列")
        print("\n列名列表:")
        for i, header in enumerate(headers, 1):
            print(f"  列{i}: {header}")

    print("\n" + "=" * 80)
    print("分析完成")
    print("=" * 80)

except ImportError as e:
    print(f"缺少依赖库: {e}")
    print("请安装: pip install openpyxl")
except Exception as e:
    print(f"发生错误: {e}")
    import traceback
    traceback.print_exc()
