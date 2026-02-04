#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
import io

# 设置标准输出为UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    from openpyxl import load_workbook

    # 读取Excel文件
    file_path = r'C:\Users\HP\Desktop\tools\18.保险丝线径匹配评审\inline.xlsx'
    wb = load_workbook(file_path, data_only=True)

    print("=" * 80)
    print("Excel File Analysis Report")
    print("=" * 80)

    # 显示所有工作表
    print(f"\nWorksheet names: {wb.sheetnames}")

    # 读取第一个工作表
    ws = wb.active
    print(f"\nCurrent worksheet: {ws.title}")
    print(f"Data range: {ws.max_row} rows x {ws.max_column} columns")

    # 读取所有数据
    data = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        data.append(row)

    # 分析列名
    print("\n" + "=" * 80)
    print("Column Analysis:")
    print("=" * 80)

    if len(data) > 0:
        headers = data[0]
        print(f"\nTotal columns: {len(headers)}")
        print("\nColumn names:")
        for i, header in enumerate(headers, 1):
            print(f"  Column {i}: {header}")

    # 数据预览
    print("\n" + "=" * 80)
    print("Data Preview (First 10 rows):")
    print("=" * 80)

    for row_idx in range(min(11, len(data))):
        print(f"\nRow {row_idx + 1}:")
        row = data[row_idx]
        for col_idx, cell_value in enumerate(row, 1):
            print(f"  Column {col_idx}: {cell_value}")

    # 统计信息
    print("\n" + "=" * 80)
    print("Statistics:")
    print("=" * 80)
    print(f"Total data rows (excluding header): {len(data) - 1}")
    print(f"Total columns: {len(headers) if len(data) > 0 else 0}")

    print("\n" + "=" * 80)
    print("Analysis Complete")
    print("=" * 80)

except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Please install: pip install openpyxl")
except Exception as e:
    print(f"Error occurred: {e}")
    import traceback
    traceback.print_exc()
