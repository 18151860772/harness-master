#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
inline.xlsx 文件详细分析报告
"""

import sys
import io
from openpyxl import load_workbook

# 设置标准输出为UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_inline_file():
    file_path = r'C:\Users\HP\Desktop\tools\18.保险丝线径匹配评审\inline.xlsx'

    # 读取Excel文件
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # 读取所有数据
    data = list(ws.iter_rows(values_only=True))

    print("=" * 100)
    print(" " * 35 + "INLINE.XLSX 文件结构分析报告")
    print("=" * 100)

    # 1. 基本信息
    print("\n【一、文件基本信息】")
    print("-" * 100)
    print(f"工作表名称: {wb.sheetnames[0]}")
    print(f"数据范围: {ws.max_row} 行 x {ws.max_column} 列")
    print(f"数据行数(不含表头): {len(data) - 1} 行")

    # 2. 列信息
    print("\n【二、列结构】")
    print("-" * 100)
    print(f"总列数: {len(data[0])} 列")
    print("\n列详细信息:")
    print(f"  第1列 - 列名: '{data[0][0]}'")
    print(f"          含义: INLINE连接器左侧端子/保险丝编号")
    print(f"          数据类型: 文本")
    print(f"\n  第2列 - 列名: '{data[0][1]}'")
    print(f"          含义: INLINE连接器右侧端子/保险丝编号")
    print(f"          数据类型: 文本")

    # 3. 命名规则分析
    print("\n【三、命名规则分析】")
    print("-" * 100)
    print("\nLEFT列(INLINE-LEFT)命名特点:")
    print("  - 格式: 4-6位字母数字组合")
    print("  - 前缀模式: 主要以'BD'开头(Board/Connector)")
    print("  - 常见前缀: BD, EG, IP, FB, FL, FR, TR, RB, RM, RP等")
    print("  - 编号示例: BDBT, BDDHT, BDDP1, BDEG5, BDIPM_ETHK")
    print("  - 特殊标记: 部分带有下划线和功能标识(如_ETHK, _ANT, _GNSS)")

    print("\nRIGHT列(INLINE-RIGHT)命名特点:")
    print("  - 大多数是LEFT列的镜像或对应关系")
    print("  - 字母重组规律: 将LEFT中的字母重新排列")
    print("  - 示例对应关系:")
    print("      BDBT  ->  BTBD  (BD和BT互换位置)")
    print("      BDDHT ->  DHTBD (字母重组)")
    print("      EGBD5 ->  BDEG5 (BD和EG互换)")

    # 4. 前5行数据示例
    print("\n【四、前5行数据示例】")
    print("-" * 100)
    print(f"{'序号':<6} {'INLINE-LEFT':<25} {'INLINE-RIGHT':<25}")
    print("-" * 100)
    for i in range(1, 6):
        left_val = data[i][0] if data[i][0] else ''
        right_val = data[i][1] if data[i][1] else ''
        print(f"{i:<6} {left_val:<25} {right_val:<25}")

    # 5. 数据统计
    print("\n【五、数据统计】")
    print("-" * 100)

    left_values = [row[0] for row in data[1:]]
    right_values = [row[1] for row in data[1:]]

    # 统计前缀
    left_prefixes = {}
    for val in left_values:
        prefix = val[:2] if len(val) >= 2 else val
        left_prefixes[prefix] = left_prefixes.get(prefix, 0) + 1

    right_prefixes = {}
    for val in right_values:
        prefix = val[:2] if len(val) >= 2 else val
        right_prefixes[prefix] = right_prefixes.get(prefix, 0) + 1

    print("\nLEFT列前缀分布:")
    for prefix, count in sorted(left_prefixes.items(), key=lambda x: x[1], reverse=True):
        print(f"  {prefix}: {count} 个")

    print("\nRIGHT列前缀分布:")
    for prefix, count in sorted(right_prefixes.items(), key=lambda x: x[1], reverse=True):
        print(f"  {prefix}: {count} 个")

    # 6. 特殊模式
    print("\n【六、特殊模式识别】")
    print("-" * 100)

    # 查找带下划线的编号
    with_underscore = [(row[0], row[1]) for row in data[1:] if '_' in str(row[0])]
    if with_underscore:
        print(f"\n带功能标识的记录(共{len(with_underscore)}条):")
        for left, right in with_underscore[:10]:
            print(f"  {left:<25} -> {right}")

    # 查找重复的LEFT值
    left_counts = {}
    for val in left_values:
        left_counts[val] = left_counts.get(val, 0) + 1

    duplicates = {k: v for k, v in left_counts.items() if v > 1}
    if duplicates:
        print(f"\nLEFT列重复值(共{len(duplicates)}组):")
        for val, count in duplicates.items():
            print(f"  {val}: 出现 {count} 次")

    # 查找重复的RIGHT值
    right_counts = {}
    for val in right_values:
        right_counts[val] = right_counts.get(val, 0) + 1

    duplicates_right = {k: v for k, v in right_counts.items() if v > 1}
    if duplicates_right:
        print(f"\nRIGHT列重复值(共{len(duplicates_right)}组):")
        for val, count in list(duplicates_right.items())[:10]:
            print(f"  {val}: 出现 {count} 次")

    # 7. 用途分析
    print("\n【七、文件用途分析】")
    print("-" * 100)
    print("""
    根据文件结构分析,该文件用于:
    1. INLINE保险丝/连接器的左右端子映射关系
    2. 电路设计中保险丝与线束的对应关系
    3. 可能用于汽车电气系统中的保险丝盒(Inline Fuse Box)配置

    典型应用场景:
    - 汽车线束设计
    - 电路原理图绘制
    - 保险丝选型与匹配
    - 电气系统故障诊断
    """)

    print("\n" + "=" * 100)
    print(" " * 40 + "分析完成")
    print("=" * 100)

if __name__ == "__main__":
    analyze_inline_file()
