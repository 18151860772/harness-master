#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
保险丝匹配清单Excel导出工具
从网页应用导出数据，生成带颜色和格式的Excel文件
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 颜色定义（与网页一致）
COLORS = {
    'lime': 'FF7FFF00',      # 亮绿色 - 标准匹配、保险丝代号
    'orange': 'FFFF8C00',    # 橙色 - Option、接近标准
    'cyan': 'FF00FFFF',      # 青色 - 线径相关
    'gray': 'FF999999',      # 灰色 - 次要文本
    'white': 'FFFFFFFF',     # 白色 - 背景
    'black': 'FF000000',     # 黑色 - 主要文本
    'header_gray': 'FF404040' # 表头深灰色
}

def get_cell_style(style_type):
    """
    获取单元格样式

    Args:
        style_type: 样式类型 ('header', 'fuse_code', 'wire_id', 'option',
                    'to', 'to_wire_ids', 'to_wire_options', 'to_wire_diameters',
                    'fuse_function', 'wire_diameter', 'note')
    """
    styles = {
        'header': {
            'font': Font(name='Arial', size=11, bold=True, color='FFFFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_gray'], end_color=COLORS['header_gray']),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'fuse_code': {
            'font': Font(name='Arial', size=10, bold=True, color=COLORS['lime']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top'),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'wire_id': {
            'font': Font(name='Courier New', size=10, bold=True, color=COLORS['black']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top'),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'option': {
            'font': Font(name='Arial', size=10, bold=True, color=COLORS['orange']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top'),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'from': {
            'font': Font(name='Arial', size=10, color=COLORS['black']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top'),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'to': {
            'font': Font(name='Arial', size=9, color=COLORS['black']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'to_wire_ids': {
            'font': Font(name='Courier New', size=9, color=COLORS['gray']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'to_wire_options': {
            'font': Font(name='Courier New', size=9, color=COLORS['orange']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'to_wire_diameters': {
            'font': Font(name='Courier New', size=9, color=COLORS['cyan']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'fuse_function': {
            'font': Font(name='Arial', size=9, color=COLORS['black']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'wire_diameter': {
            'font': Font(name='Courier New', size=10, color=COLORS['black']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top'),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'note': {
            'font': Font(name='Arial', size=10, color=COLORS['black']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top'),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        },
        'default': {
            'font': Font(name='Arial', size=10, color=COLORS['black']),
            'fill': PatternFill(start_color=COLORS['white'], end_color=COLORS['white']),
            'alignment': Alignment(vertical='top'),
            'border': Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
        }
    }

    return styles.get(style_type, styles['default'])

def apply_note_color(cell):
    """根据备注内容设置颜色"""
    if cell.value and isinstance(cell.value, str):
        if '标准' in cell.value:
            cell.font = Font(name='Arial', size=10, color=COLORS['lime'])
        elif '接近' in cell.value:
            cell.font = Font(name='Arial', size=10, color=COLORS['orange'])

def set_column_widths(ws):
    """设置列宽"""
    column_widths = {
        'A': 12,  # 保险丝代号
        'B': 10,  # 回路号
        'C': 8,   # Option
        'D': 12,  # FROM
        'E': 15,  # TO
        'F': 40,  # TO对应回路号（翻倍）
        'G': 30,  # 回路号Option（翻倍）
        'H': 12,  # 回路线径
        'I': 40,  # 保险丝功能
        'J': 8,   # 线径（调窄）
        'K': 10,  # 电路类型
        'L': 10,  # 线束类型
        'M': 10,  # 推荐保险丝（调窄）
        'N': 10,  # 保险丝类型（调窄）
        'O': 10   # 备注（调窄）
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

def export_to_excel(data, output_file='保险丝匹配清单.xlsx'):
    """
    导出数据到Excel文件

    Args:
        data: 数据列表（从JavaScript导出的JSON数据）
        output_file: 输出文件名
    """
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '保险丝匹配清单'

    # 设置列宽
    set_column_widths(ws)

    # 表头
    headers = [
        '保险丝代号', '回路号', 'Option', 'FROM', 'TO',
        'TO对应回路号', '回路号Option', '回路线径', '保险丝功能',
        '线径(mm²)', '电路类型', '线束类型', '推荐保险丝', '保险丝类型', '备注'
    ]

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        style = get_cell_style('header')
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

    # 写入数据
    style_mapping = [
        'fuse_code',      # A
        'wire_id',        # B
        'option',         # C
        'from',           # D
        'to',             # E
        'toWireIds',      # F
        'toWireOptions',  # G
        'toWireDiameters',# H
        'fuseFunction',   # I
        'wireDiameter',   # J
        'circuitType',    # K
        'harnessType',    # L
        'fuseRating',     # M
        'fuseType',       # N
        'note'            # O
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, field in enumerate(style_mapping, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # 获取字段值
            field_mapping = {
                'fuseCode': '保险丝代号',
                'wireId': '回路号',
                'option': 'Option',
                'from': 'FROM',
                'to': 'TO',
                'toWireIds': 'TO对应回路号',
                'toWireOptions': '回路号Option',
                'toWireDiameters': '回路线径',
                'fuseFunction': '保险丝功能',
                'wireDiameter': '线径(mm²)',
                'circuitType': '电路类型',
                'harnessType': '线束类型',
                'fuseRating': '推荐保险丝',
                'fuseType': '保险丝类型',
                'note': '备注'
            }

            # 获取值
            value = row_data.get(field, '')
            if field == 'wireDiameter' and isinstance(value, (int, float)):
                value = f"{value:.2f}"
            elif field == 'toDisplay':
                value = row_data.get('toDisplay', '')
            elif field == 'from':
                from_val = row_data.get('from', '')
                from_pin = row_data.get('fromPin', '')
                value = f"{from_val}-{from_pin}" if from_pin else from_val
            elif field == 'to':
                value = row_data.get('toDisplay', '')

            cell.value = value

            # 应用样式
            style_type = 'default'
            if field in style_mapping:
                style_type = field

            style = get_cell_style(style_type)
            cell.font = style['font']
            cell.fill = style['fill']
            cell.alignment = style['alignment']
            cell.border = style['border']

            # 备注列特殊处理颜色
            if field == 'note':
                apply_note_color(cell)

    # 冻结表头行
    ws.freeze_panes(1, 0)

    # 保存文件
    wb.save(output_file)
    print(f"✅ Excel文件已生成：{output_file}")
    print(f"📊 共导出 {len(data)} 条数据")

def main():
    """主函数：从JSON文件读取数据并导出Excel"""
    print("🚀 保险丝匹配清单Excel导出工具")
    print("=" * 50)

    # 提示用户导出JSON数据
    print("\n📝 使用方法：")
    print("1. 在网页应用中按F12打开开发者工具")
    print("2. 在Console控制台中输入：")
    print("   copy(JSON.stringify(filteredData))")
    print("3. 将复制的JSON数据保存到 filtered_data.json 文件")
    print("4. 运行此脚本")

    # 尝试读取JSON文件
    import os
    json_file = 'filtered_data.json'

    if os.path.exists(json_file):
        print(f"\n📂 读取数据文件：{json_file}")
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # 生成输出文件名
        timestamp = datetime.now().strftime('%Y%m%d')
        output_file = f'保险丝匹配清单_{timestamp}.xlsx'

        # 导出到Excel
        export_to_excel(data, output_file)

        print(f"\n💾 文件已保存到：{os.path.abspath(output_file)}")
    else:
        print(f"\n❌ 错误：未找到数据文件 {json_file}")
        print(f"   请确保文件存在于当前目录：{os.path.abspath('.')}")
        print("\n📖 示例数据格式：")
        print(json.dumps([
            {
                "fuseCode": "UEC-F10",
                "wireId": "W001",
                "option": "BASE",
                "from": "UEC",
                "fromPin": "1A",
                "to": "IEC",
                "toPin": "2B",
                "toDisplay": "IEC-2B\nE101-3",
                "toWireIds": "W002\nW003",
                "toWireOptions": "PREM\nBASE",
                "toWireDiameters": "0.75\n1.00",
                "fuseFunction": "仪表盒\n左前大灯",
                "wireDiameter": 0.5,
                "circuitType": "电源",
                "harnessType": "前部线束",
                "fuseRating": "10A",
                "fuseType": "MIN",
                "note": "标准匹配"
            }
        ], indent=2, ensure_ascii=False))

if __name__ == '__main__':
    main()
