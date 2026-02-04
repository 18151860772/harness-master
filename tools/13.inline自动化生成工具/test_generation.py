# -*- coding: utf-8 -*-
import openpyxl

# 读取最近生成的文件
import os
import glob

uploads_dir = r'C:\Users\HP\Desktop\tools\inline自动化生成工具\uploads'
files = glob.glob(os.path.join(uploads_dir, 'inline清单_*.xlsx'))

if files:
    latest_file = max(files, key=os.path.getmtime)
    print(f'检查文件: {os.path.basename(latest_file)}')

    wb = openpyxl.load_workbook(latest_file)
    ws = wb[wb.sheetnames[0]]

    print(f'\n工作表: {wb.sheetnames[0]}')

    # 检查第1行
    print('\n第1行:')
    print(f'  A1: {ws["A1"].value}')
    print(f'  L1: {ws["L1"].value}')

    # 检查第2行
    print('\n第2行:')
    print(f'  A2: {ws["A2"].value}')
    print(f'  L2: {ws["L2"].value}')

    # 检查第3行表头
    print('\n第3行表头 (左侧 B3:K3):')
    headers = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    for col in headers:
        cell = ws[f'{col}3']
        print(f'  {col}3: {cell.value}')

    print('\n第3行表头 (右侧 L3:V3):')
    headers_r = ['L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V']
    for col in headers_r:
        cell = ws[f'{col}3']
        print(f'  {col}3: {cell.value}')

    # 检查合并单元格
    print('\n合并单元格:')
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= 3:
            print(f'  {merged}')

    # 检查格式
    print('\n格式检查:')
    print(f'  A1字体: {ws["A1"].font.name}, 大小: {ws["A1"].font.size}, 加粗: {ws["A1"].font.bold}')
    print(f'  第1行行高: {ws.row_dimensions[1].height}')
    print(f'  第2行行高: {ws.row_dimensions[2].height}')
    print(f'  第3行行高: {ws.row_dimensions[3].height}')
else:
    print('未找到生成的文件')
