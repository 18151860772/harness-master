# -*- coding: utf-8 -*-
import openpyxl
import sys

try:
    # 读取Excel文件
    file_path = r'ECR-T28-0S_20260119.xlsx'
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 获取所有sheet名称
    print('=== 所有Sheet名称 ===')
    for i, sheet_name in enumerate(wb.sheetnames):
        print(f'{i+1}. {sheet_name}')

    # 查找包含Masterlist的sheet
    target_sheet_name = None
    for sheet_name in wb.sheetnames:
        if 'master' in sheet_name.lower() or 'list' in sheet_name.lower():
            target_sheet_name = sheet_name
            break

    # 如果没找到，使用最后一个sheet
    if not target_sheet_name:
        target_sheet_name = wb.sheetnames[-1]

    print(f'\n=== 分析Sheet: {target_sheet_name} ===\n')

    ws = wb[target_sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    print(f'总行数: {max_row}, 总列数: {max_col}\n')

    # 读取前10行
    print('=== 前10行数据 ===')
    for row_idx in range(1, min(11, max_row + 1)):
        row_data = []
        for col_idx in range(1, min(21, max_col + 1)):  # 只显示前20列
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is None:
                cell_value = ''
            row_data.append(str(cell_value)[:20])  # 截断过长的值
        print(f'第{row_idx}行: {row_data}')

    # 特别分析第3、4、5行
    print('\n=== 重点行分析 ===')
    print('\n第3行（配置描述）:')
    row3 = []
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=3, column=col_idx).value
        if val:
            row3.append(f'列{col_idx}: {val}')
    for item in row3[:15]:
        print(f'  {item}')

    print('\n第4行（查找ALL）:')
    row4 = []
    all_found = False
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=4, column=col_idx).value
        if val:
            row4.append(f'列{col_idx}: {val}')
            if str(val).strip() == 'ALL':
                all_found = True
                print(f'  ✓ 找到ALL在第{col_idx}列！')
    if not all_found:
        print('  ⚠ 未找到ALL列')
        for item in row4[:15]:
            print(f'  {item}')

    print('\n第5行（第一个零件数据）:')
    row5 = []
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=5, column=col_idx).value
        if val is not None:
            row5.append(f'列{col_idx}: {val}')
    for item in row5[:15]:
        print(f'  {item}')

    print('\n=== 数据结构总结 ===')
    print(f'目标Sheet: {target_sheet_name}')
    if all_found:
        print(f'✓ ALL列位置: 第4行')
    print(f'配置描述行: 第3行')
    print(f'数据起始行: 第5行')

except Exception as e:
    print(f'错误: {e}')
    import traceback
    traceback.print_exc()
